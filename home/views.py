import csv
import io
import calendar
from calendar import monthrange
from collections import defaultdict
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from django.db import IntegrityError, transaction


from django.contrib import messages
from django.db.models import Sum, F, Value, DecimalField, ExpressionWrapper, Case, When, Count
from django import forms
from django.forms import ModelForm, formset_factory
from django.http import HttpResponseBadRequest, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.conf import settings
from django.urls import reverse
from django.db.models.functions import Coalesce

from .forms import TransactionForm, CSVUploadForm, TransactionImportForm, ExpenseEditForm, ExpenseAttachmentUploadForm, WithholdingPayoutForm, IncomeEditForm, TransferEditForm, BalanceAdjustmentEditForm
from .models import (
    Expense,
    ExpenseAttachment,
    Income,
    Category,
    IncomeCategory,
    BankAccount,
    ImportBatch,
    WithholdingCategory,
    WithholdingTransaction,
    Transfer,
    BalanceAdjustment,

    # ✅ Rental / CRA additions
    RentalProperty,
    RentalUnit,
    CRARentalExpenseCategory,
    PropertyMortgage,

    # ✅ Month-end close + category snapshots
    MonthEndClose,
    MonthEndExpenseCategorySnapshot,
    MonthEndWithholdingCategorySnapshot,
    MonthEndIncomeCategorySnapshot,
)


TransactionImportFormSet = formset_factory(TransactionImportForm, extra=0)

# --- Auto-mapping rules ---

EXPLICIT_EXPENSE_KEYWORD_CATEGORY_NAMES = {
    "GORE MUTUAL": "Subaru Insurance",
    "TD INS": "Arnprior Insurance",
    "FIDO MOBILE": "Cell Phone",
    "BELL CANADA": "Arnprior Internet",
    "LOBLAWS": "Groceries",
    "FOODLAND": "Groceries",
    "COSTCO": "Groceries",
    "ULTRAMAR": "Gas",
    "TIM HORTONS": "Restaurants",
    "STARBUCKS": "Restaurants",
    "TST-HUNTERS": "Restaurants",
    "WAL-MART": "Groceries",
    "PIZZA": "Restaurants",
    "MCDONALD": "Restaurants",
    "NO GO COFFEE": "Restaurants",
    "SHAWARMA": "Restaurants",
    "STINSON": "Foxview Heat",      # covers "STINSON AND SON"
    "FAT LES": "Restaurants",
    "ENBRIDGE": "Arnprior Heat",

    # New mappings requested
    "NETFLIX": "Digital Subscriptions",
    "CHATGPT": "Digital Subscriptions",
    "SPOTIFY": "Digital Subscriptions",
    "NOTION": "Digital Subscriptions",
    "CARLETON": "Golf",
    "FARM BOY": "Groceries",
    "AMAZON": "Miscellaneous",
}

EXACT_ETFR_SNOW_REMOVAL_AMOUNT = Decimal("180.80")


def get_category_cached(name, cache, missing_set):
    if not name:
        return None

    if name in cache:
        return cache[name]
    if name in missing_set:
        return None

    try:
        cat = Category.objects.get(name=name)
        cache[name] = cat
        return cat
    except Category.DoesNotExist:
        missing_set.add(name)
        return None


def apply_income_rules(desc_upper, amount, entry_type_default, parsed_date=None):
    """
    Income inference rules used only during import parsing.

    Important safety constraints:
    - Only runs when entry_type_default is already 'income' (deposit column)
    - E-TRANSFER heuristic triggers only when 'E-TRANSFER' is in description
    - GLOBALIZATION is always Employment Income
    """
    entry_type = entry_type_default
    income_source = ""

    # Employment income (explicit)
    if "GLOBALIZATION" in desc_upper:
        return "income", "Employment Income"

    # Only infer rental income on deposits that contain E-TRANSFER (substring match; unique codes ok)
    if "E-TRANSFER" in desc_upper and entry_type_default == "income":
        # If we have a date, enforce "around the 1st of month" window (± ~1.5 weeks)
        in_window = True
        if parsed_date:
            first_of_month = date(parsed_date.year, parsed_date.month, 1)
            in_window = abs((parsed_date - first_of_month).days) <= 11  # ~1.5 weeks

        if in_window:
            # Amount heuristics (±5%)
            main_target = Decimal("2500")
            loft_target = Decimal("1600")

            main_low = main_target * Decimal("0.95")
            main_high = main_target * Decimal("1.05")

            loft_low = loft_target * Decimal("0.95")
            loft_high = loft_target * Decimal("1.05")

            if main_low <= amount <= main_high:
                return "income", "Arnprior Rental Income (MAIN)"
            if loft_low <= amount <= loft_high:
                return "income", "Arnprior Rental Income (LOFT)"

        # Fallback (your existing loose thresholds) — still only for E-TRANSFER deposits
        if Decimal("2000") <= amount <= Decimal("2700"):
            return "income", "Arnprior Rental Income (MAIN)"
        elif amount < Decimal("2000"):
            return "income", "Arnprior Rental Income (LOFT)"

    return entry_type, income_source

def apply_expense_rules(desc_upper, amount, category_cache, missing_categories):
    if "E-TFR" in desc_upper and amount == EXACT_ETFR_SNOW_REMOVAL_AMOUNT:
        return get_category_cached("Arnprior Snow Removal", category_cache, missing_categories)

    for keyword, cat_name in EXPLICIT_EXPENSE_KEYWORD_CATEGORY_NAMES.items():
        if keyword in desc_upper:
            cat = get_category_cached(cat_name, category_cache, missing_categories)
            if cat:
                return cat

    return None

def get_arnprior_shared_unit_id():
    """
    Best-effort lookup for the Arnprior shared/common unit.
    Returns RentalUnit.id or None (never raises).
    SQLite-safe: uses icontains instead of regex.
    """
    qs = RentalUnit.objects.select_related("property").filter(property__name__iexact="Arnprior")
    unit = (
        qs.filter(name__icontains="shared").order_by("name").first()
        or qs.filter(name__icontains="common").order_by("name").first()
    )
    return unit.id if unit else None

def get_foxview_shared_unit_id():
    """
    Best-effort lookup for the Foxview shared/common unit.
    Returns RentalUnit.id or None (never raises).
    """
    qs = RentalUnit.objects.select_related("property").filter(property__name__iexact="Foxview")
    unit = (
        qs.filter(name__icontains="shared").order_by("name").first()
        or qs.filter(name__icontains="common").order_by("name").first()
    )
    return unit.id if unit else None

def build_income_rental_unit_map():
    """
    Used by import review UI to display inferred rental unit for an IncomeCategory.
    Returns: { "<income_category_id>": "Property — Unit" }
    """
    m = {}
    for ic in IncomeCategory.objects.select_related("default_rental_unit__property").all():
        if ic.default_rental_unit_id:
            m[str(ic.id)] = f"{ic.default_rental_unit.property.name} — {ic.default_rental_unit.name}"
    return m

def dashboard(request):
    today = date.today()
    selected_month_str = request.GET.get("month", today.strftime("%Y-%m"))

    try:
        year, month = map(int, selected_month_str.split("-"))
    except ValueError:
        year, month = today.year, today.month

    selected_date = date(year, month, 1)
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    selected_month_display = first_day.strftime("%B %Y")

    # Month navigation
    prev_month_date = date(year, month, 1) - timedelta(days=1)  # Go to last day of previous month
    prev_month = prev_month_date.strftime("%Y-%m")

    current_month = today.strftime("%Y-%m")

    # Next month calculation
    if month == 12:
        next_month_date = date(year + 1, 1, 1)
    else:
        next_month_date = date(year, month + 1, 1)
    next_month = next_month_date.strftime("%Y-%m")

    # Check if this is an AJAX request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == "POST":
        # -------------------------
        # Expense modal edit/delete
        # -------------------------
        if "expense_id" in request.POST:
            expense = get_object_or_404(Expense, pk=request.POST["expense_id"])
            if "delete_expense" in request.POST:
                expense_id = expense.id
                expense.delete()

                if is_ajax:
                    return JsonResponse({
                        'deleted': True,
                        'transaction_id': expense_id
                    })

                selected_month_param = f"{today.year:04d}-{today.month:02d}"
                return redirect(f"/?month={selected_month_param}")
            else:
                try:
                    expense.date = datetime.strptime(request.POST["date"], "%Y-%m-%d").date()
                    expense.vendor_name = request.POST["vendor_name"]

                    category_name = request.POST["category"]
                    expense.category = get_object_or_404(Category, name=category_name)

                    expense.location = request.POST.get("location", "Ottawa")

                    amount_str = (request.POST.get("amount") or "").strip()
                    if amount_str:
                        try:
                            expense.amount = Decimal(amount_str)
                        except (InvalidOperation, ValueError):
                            if is_ajax:
                                return JsonResponse({
                                    'error': 'Invalid amount format'
                                }, status=400)
                            pass

                    expense.notes = request.POST.get("notes", "")

                    # Bank account
                    bank_account_id = (request.POST.get("bank_account") or "").strip()
                    if bank_account_id:
                        expense.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                    else:
                        expense.bank_account = None

                    # Rental + CRA fields
                    rental_unit_id = (request.POST.get("rental_unit") or "").strip()
                    if rental_unit_id:
                        expense.rental_unit = get_object_or_404(RentalUnit, pk=rental_unit_id)
                    else:
                        expense.rental_unit = None

                    cra_category_id = (request.POST.get("cra_category") or "").strip()
                    if cra_category_id:
                        expense.cra_category = get_object_or_404(
                            CRARentalExpenseCategory,
                            pk=cra_category_id
                        )
                    else:
                        expense.cra_category = None

                    pct_str = (request.POST.get("rental_business_use_pct") or "").strip()
                    if pct_str:
                        expense.rental_business_use_pct = Decimal(pct_str)
                    else:
                        expense.rental_business_use_pct = None

                    expense.save()

                    if is_ajax:
                        # Format rental unit display
                        rental_unit_display = None
                        if expense.rental_unit:
                            rental_unit_display = f"{expense.rental_unit.property.name} — {expense.rental_unit.name}"

                        return JsonResponse({
                            'success': True,
                            'transaction': {
                                'id': expense.id,
                                'date': expense.date.strftime('%Y-%m-%d'),
                                'date_display': expense.date.strftime('%Y-%m-%d'),
                                'vendor_name': expense.vendor_name,
                                'category_name': expense.category.name,
                                'location': expense.location,
                                'amount': str(expense.amount),
                                'amount_display': f"{expense.amount:,.2f}",
                                'notes': expense.notes or '',
                                'bank_account_id': expense.bank_account.id if expense.bank_account else '',
                                'bank_account_name': expense.bank_account.name if expense.bank_account else '',
                                'rental_unit_id': expense.rental_unit.id if expense.rental_unit else '',
                                'rental_unit_display': rental_unit_display or '',
                                'cra_category_id': expense.cra_category.id if expense.cra_category else '',
                                'rental_business_use_pct': str(expense.rental_business_use_pct) if expense.rental_business_use_pct else ''
                            }
                        })
                except Exception as e:
                    if is_ajax:
                        return JsonResponse({
                            'error': f'Error updating expense: {str(e)}'
                        }, status=400)
                    raise

            selected_month_param = f"{expense.date.year:04d}-{expense.date.month:02d}"
            return redirect(f"/?month={selected_month_param}")

        # -------------------------
        # Income modal edit/delete
        # -------------------------
        elif "income_id" in request.POST:
            income = get_object_or_404(Income, pk=request.POST["income_id"])
            if "delete_income" in request.POST:
                income_id = income.id
                income.delete()

                if is_ajax:
                    return JsonResponse({
                        'deleted': True,
                        'transaction_id': income_id
                    })

                selected_month_param = f"{today.year:04d}-{today.month:02d}"
                return redirect(f"/?month={selected_month_param}")
            else:
                try:
                    income.date = datetime.strptime(request.POST["date"], "%Y-%m-%d").date()

                    source_id = request.POST.get("source")
                    income_cat = get_object_or_404(IncomeCategory, pk=source_id) if source_id else None

                    income.income_category = income_cat
                    if income_cat:
                        # legacy sync for now
                        income.category = income_cat.name

                    amount_str = (request.POST.get("amount") or "").strip()
                    if amount_str:
                        try:
                            income.amount = Decimal(amount_str)
                        except (InvalidOperation, ValueError):
                            if is_ajax:
                                return JsonResponse({
                                    'error': 'Invalid amount format'
                                }, status=400)
                            pass

                    taxable = request.POST.get("taxable") == "1"
                    income.taxable = taxable

                    income.notes = request.POST.get("notes", "")

                    rental_unit_id = (request.POST.get("rental_unit") or "").strip()
                    if rental_unit_id:
                        income.rental_unit = get_object_or_404(RentalUnit, pk=rental_unit_id)
                    else:
                        income.rental_unit = None

                    bank_account_id = (request.POST.get("bank_account") or "").strip()
                    if bank_account_id:
                        income.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                    else:
                        income.bank_account = None

                    income.save()

                    if is_ajax:
                        # Format rental unit display
                        rental_unit_display = None
                        if income.rental_unit:
                            rental_unit_display = f"{income.rental_unit.property.name} — {income.rental_unit.name}"

                        return JsonResponse({
                            'success': True,
                            'transaction': {
                                'id': income.id,
                                'date': income.date.strftime('%Y-%m-%d'),
                                'date_display': income.date.strftime('%Y-%m-%d'),
                                'source_name': income.income_category.name if income.income_category else (income.category or 'Uncategorized'),
                                'income_category_id': income.income_category.id if income.income_category else '',
                                'amount': str(income.amount),
                                'amount_display': f"{income.amount:,.2f}",
                                'taxable': income.taxable,
                                'notes': income.notes or '',
                                'bank_account_id': income.bank_account.id if income.bank_account else '',
                                'bank_account_name': income.bank_account.name if income.bank_account else '',
                                'rental_unit_id': income.rental_unit.id if income.rental_unit else '',
                                'rental_unit_display': rental_unit_display or ''
                            }
                        })
                except Exception as e:
                    if is_ajax:
                        return JsonResponse({
                            'error': f'Error updating income: {str(e)}'
                        }, status=400)
                    raise

            selected_month_param = f"{income.date.year:04d}-{income.date.month:02d}"
            return redirect(f"/?month={selected_month_param}")

        # -------------------------
        # Transfer modal edit/delete
        # -------------------------
        elif "transfer_id" in request.POST:
            return handle_transfer_edit(request, is_ajax)

        # -------------------------
        # New transaction (Add Transaction form)
        # -------------------------
        else:
            form = TransactionForm(request.POST)
            if form.is_valid():
                entry_type = form.cleaned_data["entry_type"]
                entry_date = form.cleaned_data["date"]

                if entry_type == "expense":
                    expense = Expense(
                        date=entry_date,
                        vendor_name=form.cleaned_data["vendor_name"],
                        category=form.cleaned_data["category"],
                        amount=form.cleaned_data["amount"],
                        location=form.cleaned_data.get("location") or "Ottawa",
                        notes=form.cleaned_data.get("notes") or "",
                        bank_account=form.cleaned_data.get("bank_account"),
                    )

                    # Rental + CRA extras
                    expense.rental_unit = form.cleaned_data.get("rental_unit") or None
                    expense.cra_category = form.cleaned_data.get("cra_category") or None
                    pct = form.cleaned_data.get("rental_business_use_pct")
                    expense.rental_business_use_pct = pct if pct is not None else None

                    # Withholding application from expense
                    if form.cleaned_data.get("apply_to_withholding"):
                        expense.withholding_category = form.cleaned_data.get("withholding_category")
                    else:
                        expense.withholding_category = None

                    expense.save()
                    messages.success(request, "Expense saved successfully!")

                elif entry_type == "income":
                    income = Income(
                        date=entry_date,
                        income_category=form.cleaned_data.get("source"),
                        amount=form.cleaned_data["amount"],
                        taxable=form.cleaned_data.get("taxable", False),
                        notes=form.cleaned_data.get("notes") or "",
                        bank_account=form.cleaned_data.get("bank_account"),
                    )

                    if income.income_category:
                        income.category = income.income_category.name

                    income.rental_unit = form.cleaned_data.get("income_rental_unit") or None

                    income.save()
                    messages.success(request, "Income saved successfully!")

                elif entry_type == "transfer":
                    transfer = Transfer(
                        date=entry_date,
                        from_account=form.cleaned_data.get("from_account"),
                        to_account=form.cleaned_data.get("to_account"),
                        amount=form.cleaned_data["amount"],
                        notes=form.cleaned_data.get("notes") or "",
                        withholding_category=form.cleaned_data.get("withholding_category") or None,
                    )
                    transfer.save()
                    messages.success(request, "Transfer saved successfully!")

                # Only redirect if form was valid and transaction was saved
                return redirect(f"/?month={selected_month_str}")
            else:
                # Form is invalid - show errors
                messages.error(request, f"Form validation failed: {form.errors}")
    else:
        form = TransactionForm(initial={"date": selected_date})

    # -------------------------
    # Query transactions for month
    # -------------------------
    income_entries = (
        Income.objects
        .filter(date__range=(first_day, last_day))
        .select_related("income_category", "bank_account")
    )

    expense_entries = (
        Expense.objects
        .filter(date__range=(first_day, last_day))
        .select_related("category", "bank_account", "rental_unit", "cra_category")
        .annotate(attachment_count=Count("attachments", distinct=True))
    )

    # Transfers for this month (exclude split children, only show parents and non-split transfers)
    transfer_entries = (
        Transfer.objects
        .filter(date__range=(first_day, last_day))
        .filter(parent_transfer__isnull=True)  # Exclude split children
        .select_related("from_account", "to_account", "withholding_category")
        .prefetch_related("splits")  # Eager load children for display
    )

    # Exclude Business Reimbursement from totals
    total_income = sum(
        i.amount for i in income_entries
        if not i.income_category or i.income_category.name != 'Business Reimbursement'
    )
    total_expenses = sum(e.amount for e in expense_entries)
    net_savings = total_income - total_expenses
    total_transfers = sum(t.amount for t in transfer_entries)

    income_by_source = defaultdict(list)
    for income in income_entries:
        label = income.income_category.name if income.income_category else (income.category or "Uncategorized")
        income_by_source[label].append(income)

    expenses_by_category = defaultdict(Decimal)
    targets_by_category = {}

    for expense in expense_entries:
        category_name = expense.category.name
        expenses_by_category[category_name] += expense.amount

        if category_name not in targets_by_category:
            targets_by_category[category_name] = expense.category.monthly_limit or Decimal("0")

    category_summaries = []
    for cat_name, spent in expenses_by_category.items():
        target = targets_by_category.get(cat_name, Decimal("0"))
        category_summaries.append({
            "name": cat_name,
            "spent": spent,
            "target": target,
        })

    category_summaries.sort(key=lambda cs: cs["name"])

    accounts = BankAccount.objects.all().order_by("institution", "name")

    context = {
        "form": form,
        "selected_month": f"{year:04d}-{month:02d}",
        "selected_month_display": selected_month_display,
        "prev_month": prev_month,
        "current_month": current_month,
        "next_month": next_month,

        "income_entries": income_entries,
        "expense_entries": expense_entries,
        "transfer_entries": transfer_entries,

        "total_income": total_income,
        "total_expenses": total_expenses,
        "net_savings": net_savings,
        "total_transfers": total_transfers,
        "income_by_source": dict(income_by_source),
        "expenses_by_category": dict(expenses_by_category),
        "targets_by_category": targets_by_category,
        "all_categories": Category.objects.filter(is_archived=False),
        "category_summaries": category_summaries,
        "accounts": accounts,
        "income_categories": IncomeCategory.objects.all().order_by("name"),

        # Rental / CRA dropdowns for modals + add-transaction form
        "all_rental_units": RentalUnit.objects.select_related("property").order_by("property__name", "name"),
        "cra_categories": CRARentalExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name"),
        "income_source_default_unit_map": {
            str(c.id): (c.default_rental_unit_id or "")
            for c in IncomeCategory.objects.all()
        },
        "income_source_taxable_map": {
            str(c.id): c.taxable_default
            for c in IncomeCategory.objects.all()
        },
        "withholding_categories": WithholdingCategory.objects.select_related("account").order_by("account__name", "name"),
    }

    return render(request, "dashboard.html", context)




def category_progress(request):
    today = date.today()
    selected_month_str = request.GET.get("month", today.strftime("%Y-%m"))
    try:
        year, month = map(int, selected_month_str.split("-"))
    except ValueError:
        year, month = today.year, today.month

    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    selected_month_display = first_day.strftime("%B %Y")

    def get_progress_color(percent):
        """Get a granular color based on percentage (0-100+)"""
        if percent >= 100:
            return "#dc3545"  # Red - over budget
        elif percent >= 90:
            return "#fd7e14"  # Orange - approaching limit
        elif percent >= 80:
            return "#ffc107"  # Yellow - caution zone
        elif percent >= 70:
            return "#20c997"  # Teal - good progress
        elif percent >= 50:
            return "#0dcaf0"  # Cyan - moderate progress
        else:
            return "#198754"  # Green - well under budget

    def get_income_progress_color(percent):
        """Get a granular color for income (inverse logic - higher is better)"""
        if percent >= 100:
            return "#198754"  # Green - target met
        elif percent >= 90:
            return "#20c997"  # Teal - almost there
        elif percent >= 75:
            return "#0dcaf0"  # Cyan - good progress
        elif percent >= 50:
            return "#ffc107"  # Yellow - halfway
        elif percent >= 25:
            return "#fd7e14"  # Orange - behind
        else:
            return "#dc3545"  # Red - significantly behind

    # Check if the selected month is a locked close — use snapshots when available
    month_close_record = MonthEndClose.objects.filter(month=first_day, is_locked=True).first()

    # Pre-compute withholding-funded expense totals per category.
    # These represent spending from pre-saved buckets, not current cash outflows,
    # so they must be excluded from the cash flow health calculation.
    _wf_funded_by_cat = {
        row['category_id']: row['total']
        for row in Expense.objects.filter(
            date__range=(first_day, last_day),
            withholding_category__isnull=False,
        ).values('category_id').annotate(total=Sum('amount'))
    }

    # ========= EXPENSE PROGRESS =========
    expense_summaries = []
    total_expenses_actual = Decimal("0.00")
    total_expenses_remaining = Decimal("0.00")

    _expense_snaps = (
        month_close_record.expense_snapshots.select_related('category').all()
        if month_close_record else None
    )
    if _expense_snaps and _expense_snaps.exists():
        # Closed month: use frozen snapshot values
        for snap in _expense_snaps:
            monthly_limit = snap.monthly_limit
            total_spent = snap.actual_spent
            percent_used = float(total_spent / monthly_limit * 100) if monthly_limit > 0 else 0.0
            if snap.category.name != "Business Expense":
                wf_funded = _wf_funded_by_cat.get(snap.category_id, Decimal("0.00"))
                cash_flow_spent = total_spent - wf_funded
                total_expenses_actual += cash_flow_spent
                if cash_flow_spent < monthly_limit:
                    total_expenses_remaining += (monthly_limit - cash_flow_spent)
            expense_summaries.append({
                "name": snap.category.name,
                "target": monthly_limit,
                "actual": total_spent,
                "percent": round(percent_used, 1),
                "bar_color": get_progress_color(percent_used),
            })
    else:
        # Open/current month: use live model values
        for category in Category.objects.filter(is_archived=False).exclude(monthly_limit__isnull=True):
            total_spent = (
                Expense.objects.filter(category=category, date__range=(first_day, last_day))
                .aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )
            percent_used = (
                total_spent / category.monthly_limit * 100
                if category.monthly_limit and category.monthly_limit > 0
                else 0
            )

            # Exclude "Business Expense" from cash flow totals (gets reimbursed)
            if category.name != "Business Expense":
                wf_funded = _wf_funded_by_cat.get(category.id, Decimal("0.00"))
                cash_flow_spent = total_spent - wf_funded
                total_expenses_actual += cash_flow_spent
                if cash_flow_spent < category.monthly_limit:
                    total_expenses_remaining += (category.monthly_limit - cash_flow_spent)
            else:
                if total_spent < category.monthly_limit:
                    pass  # Don't add to total_expenses_remaining

            expense_summaries.append({
                "name": category.name,
                "target": category.monthly_limit,
                "actual": total_spent,
                "percent": round(percent_used, 1),
                "bar_color": get_progress_color(percent_used),
            })

    # ========= INCOME PROGRESS =========
    income_summaries = []
    total_income_actual = Decimal("0.00")
    total_income_remaining = Decimal("0.00")

    _income_snaps = (
        month_close_record.income_snapshots.select_related('income_category').all()
        if month_close_record else None
    )
    if _income_snaps and _income_snaps.exists():
        # Closed month: use frozen snapshot values
        for snap in _income_snaps:
            target = snap.monthly_target
            total_received = snap.actual_received
            percent_received = float(total_received / target * 100) if target > 0 else 0.0
            if snap.income_category.name != "Business Reimbursement":
                total_income_actual += total_received
                if target > 0 and total_received < target:
                    total_income_remaining += (target - total_received)
            income_summaries.append({
                "id": snap.income_category.id,
                "name": snap.income_category.name,
                "target": target,
                "actual": total_received,
                "percent": round(percent_received, 1),
                "bar_color": get_income_progress_color(percent_received) if target > 0 else "#6c757d",
            })
    else:
        # Open/current month: use live model values
        for inc_cat in IncomeCategory.objects.all():
            target = inc_cat.monthly_target or Decimal("0.00")
            total_received = (
                Income.objects.filter(income_category=inc_cat, date__range=(first_day, last_day))
                .aggregate(total=Sum("amount"))["total"]
                or Decimal("0.00")
            )
            if inc_cat.name != "Business Reimbursement":
                total_income_actual += total_received
                if target > 0 and total_received < target:
                    total_income_remaining += (target - total_received)
            percent_received = (total_received / target * 100) if target > 0 else 0
            income_summaries.append({
                "id": inc_cat.id,
                "name": inc_cat.name,
                "target": target,
                "actual": total_received,
                "percent": round(percent_received, 1),
                "bar_color": get_income_progress_color(percent_received) if target > 0 else "#6c757d",
            })

    income_summaries.sort(key=lambda x: x["name"].lower())
    expense_summaries.sort(key=lambda x: x["name"].lower())

    # ========= TRACK JENNA TRANSFERS AS INCOME =========
    jenna_transfers = Decimal("0.00")
    try:
        jenna_account = BankAccount.objects.get(name="Jenna (EXT)")
        # Outflows from Jenna's account = money she's sending
        jenna_transfers = (
            Transfer.objects.filter(
                date__range=(first_day, last_day),
                from_account=jenna_account
            )
            .aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )
    except BankAccount.DoesNotExist:
        pass  # Account doesn't exist, leave at 0

    # ========= WITHHOLDING / TRANSFER PROGRESS (monthly-based) =========

    buckets = WithholdingCategory.objects.select_related("account").order_by("name")

    # Track contributions and payouts separately for the selected month
    month_contribs = defaultdict(Decimal)
    month_payouts = defaultdict(Decimal)

    # Transfers linked to a bucket
    month_transfers = (
        Transfer.objects.filter(
            date__range=(first_day, last_day),
            withholding_category__isnull=False,
        )
        .select_related("withholding_category", "from_account", "to_account")
    )

    for t in month_transfers:
        bucket = t.withholding_category
        if not bucket or not bucket.account:
            continue

        # Money going INTO the bucket's account = contribution
        if t.to_account_id == bucket.account_id:
            month_contribs[bucket.id] += t.amount
        # Money leaving the bucket's account = payout
        elif t.from_account_id == bucket.account_id:
            month_payouts[bucket.id] += t.amount

    # Expenses funded from a withholding bucket
    month_bucket_expenses = (
        Expense.objects.filter(
            date__range=(first_day, last_day),
            withholding_category__isnull=False,
        ).select_related("withholding_category")
    )

    for e in month_bucket_expenses:
        bucket = e.withholding_category
        if not bucket:
            continue
        # Expense from a bucket always reduces it
        month_payouts[bucket.id] += e.amount

    # Define which buckets are "true savings" vs "future expenses"
    SAVINGS_BUCKETS = ["vacation", "wedding", "rrsp"]  # Case-insensitive matching

    # Build target overrides from snapshots for closed months
    _withholding_snaps = (
        month_close_record.withholding_snapshots.all()
        if month_close_record else None
    )
    withholding_target_overrides = (
        {snap.withholding_category_id: snap.monthly_target for snap in _withholding_snaps}
        if _withholding_snaps and _withholding_snaps.exists()
        else None
    )

    withholding_summaries = []
    total_withholding_remaining = Decimal("0.00")
    total_savings_remaining = Decimal("0.00")
    total_withholding_actual = Decimal("0.00")
    total_savings_actual = Decimal("0.00")

    for bucket in buckets:
        if withholding_target_overrides is not None:
            # Closed month: only show buckets active at close time; use snapshot target
            if bucket.id not in withholding_target_overrides:
                continue
            monthly_target = withholding_target_overrides[bucket.id]
        else:
            monthly_target = bucket.monthly_target or Decimal("0.00")

        # Only show buckets with a monthly target > 0
        if monthly_target <= 0:
            continue

        contrib = month_contribs.get(bucket.id, Decimal("0.00"))
        payout = month_payouts.get(bucket.id, Decimal("0.00"))
        net = contrib - payout

        balance = bucket.balance                # legacy ledger still used for now for balance
        overall_target = bucket.target_amount   # yearly/overall target
        remaining = bucket.remaining_to_target()

        # Check if this is a savings bucket
        is_savings = any(savings_name.lower() in bucket.name.lower() for savings_name in SAVINGS_BUCKETS)

        # Track realized contributions (money already set aside this month)
        if is_savings:
            total_savings_actual += contrib
        else:
            total_withholding_actual += contrib

        # Calculate remaining planned contributions
        remaining_contrib = (monthly_target - contrib) if contrib < monthly_target else Decimal("0.00")

        if is_savings:
            total_savings_remaining += remaining_contrib
        else:
            total_withholding_remaining += remaining_contrib

        # Progress bar is based on monthly contributions vs monthly_target
        percent = (contrib / monthly_target) * Decimal("100")
        percent = float(round(percent, 1))

        withholding_summaries.append(
            {
                "id": bucket.id,
                "name": bucket.name,
                "balance": balance,
                "overall_target": overall_target,
                "remaining": remaining,
                "monthly_target": monthly_target,
                "month_contrib": contrib,
                "month_payout": payout,
                "month_net": net,
                "percent": percent,
                "bar_color": get_income_progress_color(percent),  # Use income color logic for contributions
                "is_savings": is_savings,
            }
        )

    # ========= CASH FLOW HEALTH CALCULATION =========
    # Calculate total projected income
    total_income_with_jenna = total_income_actual + jenna_transfers
    total_projected_income = total_income_with_jenna + total_income_remaining

    # Calculate all outflows (actual and planned)
    # Expenses: actual spent + remaining budget
    total_expenses_projected = total_expenses_actual + total_expenses_remaining

    # Future expense buckets: money already set aside + remaining planned contributions
    total_withholding_projected = total_withholding_actual + total_withholding_remaining

    # Savings: money already saved + remaining planned contributions
    total_savings_projected = total_savings_actual + total_savings_remaining

    # Final projected balance = Income - all outflows
    cash_flow_balance = (
        total_projected_income
        - total_expenses_projected
        - total_withholding_projected
        - total_savings_projected
    )

    # Determine health status based on final balance
    if cash_flow_balance > 0:
        health_status = "positive"
        health_class = "success"
        health_icon = "✅"
    elif cash_flow_balance < 0:
        health_status = "negative"
        health_class = "danger"
        health_icon = "⚠️"
    else:
        health_status = "neutral"
        health_class = "secondary"
        health_icon = "➖"

    # Filter summaries to pinned categories if the user has any pinned
    from .models import UserProfile
    show_all = request.GET.get("show_all") == "1"
    try:
        profile = request.user.profile
        pinned_expense_names = set(profile.pinned_categories.values_list("name", flat=True))
        pinned_income_names = set(profile.pinned_income_categories.values_list("name", flat=True))
        pinned_withholding_names = set(profile.pinned_withholding_categories.values_list("name", flat=True))
    except UserProfile.DoesNotExist:
        pinned_expense_names = pinned_income_names = pinned_withholding_names = set()

    has_any_pins = any([pinned_expense_names, pinned_income_names, pinned_withholding_names])
    pinned_only = has_any_pins and not show_all
    if pinned_only:
        if pinned_expense_names:
            expense_summaries = [s for s in expense_summaries if s["name"] in pinned_expense_names]
        if pinned_income_names:
            income_summaries = [s for s in income_summaries if s["name"] in pinned_income_names]
        if pinned_withholding_names:
            withholding_summaries = [s for s in withholding_summaries if s["name"] in pinned_withholding_names]

    context = {
        "income_summaries": income_summaries,
        "expense_summaries": expense_summaries,
        "withholding_summaries": withholding_summaries,
        "selected_month": f"{year:04d}-{month:02d}",
        "selected_month_display": selected_month_display,
        "pinned_only": pinned_only,
        # Cash flow health data
        "total_income_actual": total_income_actual,
        "jenna_transfers": jenna_transfers,
        "total_income_remaining": total_income_remaining,
        "total_income_with_jenna": total_income_with_jenna,
        "total_projected_income": total_projected_income,
        "total_expenses_actual": total_expenses_actual,
        "total_expenses_remaining": total_expenses_remaining,
        "total_withholding_actual": total_withholding_actual,
        "total_withholding_remaining": total_withholding_remaining,
        "total_savings_actual": total_savings_actual,
        "total_savings_remaining": total_savings_remaining,
        "cash_flow_balance": cash_flow_balance,
        "health_status": health_status,
        "health_class": health_class,
        "health_icon": health_icon,
        "savings_buckets": SAVINGS_BUCKETS,  # For display in template
    }
    return render(request, "category_progress.html", context)

def category_expense_list(request, category_name):
    today = date.today()
    selected_range = request.GET.get("range", "12")  # Default: Last 12 months

    category = get_object_or_404(Category, name=category_name)

    # Build range options
    current_year = today.year
    range_options = [
        ("12", "Last 12 months"),
        ("ytd", f"Year to Date ({current_year})"),
    ]
    # Add previous years
    for year in range(current_year, current_year - 5, -1):
        range_options.append((str(year), str(year)))

    # Calculate date range based on selection
    if selected_range == "ytd":
        first_day = date(current_year, 1, 1)
        last_day = date(current_year, 12, 31)  # Include future transactions through end of year
    elif selected_range == "12":
        # Last 12 months
        first_day = date(today.year - 1, today.month, 1)
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
    else:
        # Specific year
        try:
            year = int(selected_range)
            first_day = date(year, 1, 1)
            last_day = date(year, 12, 31)
        except (ValueError, TypeError):
            # Fallback to last 12 months
            first_day = date(today.year - 1, today.month, 1)
            last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])

    # Generate month buckets within the date range
    def add_months(d: date, delta_months: int) -> date:
        y = d.year + (d.month - 1 + delta_months) // 12
        m = (d.month - 1 + delta_months) % 12 + 1
        return date(y, m, 1)

    month_starts = []
    cur = date(first_day.year, first_day.month, 1)
    end_month = date(last_day.year, last_day.month, 1)
    while cur <= end_month:
        month_starts.append(cur)
        cur = add_months(cur, 1)

    monthly_limit = category.monthly_limit or Decimal("0.00")
    has_limit = category.monthly_limit is not None and category.monthly_limit > 0

    month_rows = []
    trend_labels = []
    trend_values = []

    month_starts_chrono = list(reversed(month_starts))

    for m in month_starts_chrono:
        start = date(m.year, m.month, 1)
        end = date(m.year, m.month, monthrange(m.year, m.month)[1])

        expenses_qs = (
            Expense.objects
            .filter(category=category, date__range=(start, end))
            .select_related("bank_account")
            .order_by("-date", "-id")
        )

        total = expenses_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        percent = (total / monthly_limit * 100) if has_limit else 0

        month_rows.append({
            "month": m,
            "expenses": expenses_qs,
            "total": total,
            "percent": round(percent, 1),
        })

        trend_labels.append(m.strftime("%b %Y"))
        trend_values.append(float(total))

    # Reverse trend data so graph shows oldest on left, newest on right
    trend_labels = list(reversed(trend_labels))
    trend_values = list(reversed(trend_values))

    # Keep month_rows in reverse chronological order (newest first) for table display
    # (Do NOT reverse month_rows - we want newest at top)

    range_total = sum((r["total"] for r in month_rows), Decimal("0.00"))

    # Handle expense edit/delete from modal
    if request.method == "POST":
        if "expense_id" in request.POST:
            expense = get_object_or_404(Expense, pk=request.POST["expense_id"])
            if "delete_expense" in request.POST:
                expense.delete()
            else:
                expense.date = datetime.strptime(request.POST["date"], "%Y-%m-%d").date()
                expense.vendor_name = request.POST["vendor_name"]

                category_name = request.POST["category"]
                expense.category = get_object_or_404(Category, name=category_name)

                expense.location = request.POST.get("location", "Ottawa")

                amount_str = (request.POST.get("amount") or "").strip()
                if amount_str:
                    try:
                        expense.amount = Decimal(amount_str)
                    except (InvalidOperation, ValueError):
                        pass

                expense.notes = request.POST.get("notes", "")

                # Bank account
                bank_account_id = (request.POST.get("bank_account") or "").strip()
                if bank_account_id:
                    expense.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                else:
                    expense.bank_account = None

                # Rental + CRA fields
                rental_unit_id = (request.POST.get("rental_unit") or "").strip()
                if rental_unit_id:
                    expense.rental_unit = get_object_or_404(RentalUnit, pk=rental_unit_id)
                else:
                    expense.rental_unit = None

                cra_category_id = (request.POST.get("cra_category") or "").strip()
                if cra_category_id:
                    expense.cra_category = get_object_or_404(
                        CRARentalExpenseCategory,
                        pk=cra_category_id
                    )
                else:
                    expense.cra_category = None

                pct_str = (request.POST.get("rental_business_use_pct") or "").strip()
                if pct_str:
                    expense.rental_business_use_pct = Decimal(pct_str)
                else:
                    expense.rental_business_use_pct = None

                expense.save()

            return redirect("category_expense_list", category_name=category.name)

    context = {
        "category": category,
        "selected_range": selected_range,
        "range_options": range_options,
        "has_limit": has_limit,
        "monthly_limit": monthly_limit,
        "month_rows": month_rows,
        "trend_labels": trend_labels,
        "trend_values": trend_values,
        "range_total": range_total,
        # Modal dropdowns
        "all_categories": Category.objects.filter(is_archived=False).order_by("name"),
        "accounts": BankAccount.objects.all().order_by("institution", "name"),
        "all_rental_units": RentalUnit.objects.select_related("property").order_by("property__name", "name"),
        "cra_categories": CRARentalExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name"),
    }
    return render(request, "category_expense_list.html", context)

def income_category_income_list(request, pk):
    today = date.today()
    selected_range = request.GET.get("range", "12")  # Default: Last 12 months

    inc_cat = get_object_or_404(IncomeCategory, pk=pk)

    # Handle POST request for edit/delete
    if request.method == "POST":
        if "income_id" in request.POST:
            income = get_object_or_404(Income, pk=request.POST["income_id"])
            if "delete_income" in request.POST:
                income.delete()
            else:
                income.date = datetime.strptime(request.POST["date"], "%Y-%m-%d").date()

                source_id = request.POST.get("source")
                income_cat_obj = get_object_or_404(IncomeCategory, pk=source_id) if source_id else None

                income.income_category = income_cat_obj
                if income_cat_obj:
                    # legacy sync for now
                    income.category = income_cat_obj.name

                amount_str = (request.POST.get("amount") or "").strip()
                if amount_str:
                    try:
                        income.amount = Decimal(amount_str)
                    except (InvalidOperation, ValueError):
                        pass

                taxable = request.POST.get("taxable") == "1"
                income.taxable = taxable

                income.notes = request.POST.get("notes", "")

                rental_unit_id = (request.POST.get("rental_unit") or "").strip()
                if rental_unit_id:
                    income.rental_unit = get_object_or_404(RentalUnit, pk=rental_unit_id)
                else:
                    income.rental_unit = None

                bank_account_id = (request.POST.get("bank_account") or "").strip()
                if bank_account_id:
                    income.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                else:
                    income.bank_account = None

                income.save()

            # Redirect back to the same page with the same range parameter
            return redirect(f"{request.path}?range={selected_range}")

    # Build range options
    current_year = today.year
    range_options = [
        ("12", "Last 12 months"),
        ("ytd", f"Year to Date ({current_year})"),
    ]
    # Add previous years
    for year in range(current_year, current_year - 5, -1):
        range_options.append((str(year), str(year)))

    # Calculate date range based on selection
    if selected_range == "ytd":
        first_day = date(current_year, 1, 1)
        last_day = date(current_year, 12, 31)  # Include future transactions through end of year
    elif selected_range == "12":
        # Last 12 months
        first_day = date(today.year - 1, today.month, 1)
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
    else:
        # Specific year
        try:
            year = int(selected_range)
            first_day = date(year, 1, 1)
            last_day = date(year, 12, 31)
        except (ValueError, TypeError):
            # Fallback to last 12 months
            first_day = date(today.year - 1, today.month, 1)
            last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])

    # Generate month buckets within the date range
    def add_months(d: date, delta_months: int) -> date:
        y = d.year + (d.month - 1 + delta_months) // 12
        m = (d.month - 1 + delta_months) % 12 + 1
        return date(y, m, 1)

    month_starts = []
    cur = date(first_day.year, first_day.month, 1)
    end_month = date(last_day.year, last_day.month, 1)
    while cur <= end_month:
        month_starts.append(cur)
        cur = add_months(cur, 1)

    target = inc_cat.monthly_target or Decimal("0.00")

    month_rows = []
    trend_labels = []
    trend_values = []

    month_starts_chrono = list(reversed(month_starts))

    for m in month_starts_chrono:
        start = date(m.year, m.month, 1)
        end = date(m.year, m.month, monthrange(m.year, m.month)[1])

        incomes_qs = (
            Income.objects
            .filter(income_category=inc_cat, date__range=(start, end))
            .select_related("bank_account")
            .order_by("-date", "-id")
        )

        total = incomes_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        percent = (total / target * 100) if target > 0 else 0

        month_rows.append({
            "month": m,
            "incomes": incomes_qs,
            "total": total,
            "percent": round(percent, 1),
        })

        trend_labels.append(m.strftime("%b %Y"))
        trend_values.append(float(total))

    # Reverse trend data so graph shows oldest on left, newest on right
    trend_labels = list(reversed(trend_labels))
    trend_values = list(reversed(trend_values))

    # Keep month_rows in reverse chronological order (newest first) for table display
    # (Do NOT reverse month_rows - we want newest at top)

    range_total = sum((r["total"] for r in month_rows), Decimal("0.00"))

    context = {
        "income_category": inc_cat,
        "selected_range": selected_range,
        "range_options": range_options,
        "has_target": target > 0,
        "monthly_target": target,
        "month_rows": month_rows,
        "trend_labels": trend_labels,
        "trend_values": trend_values,
        "range_total": range_total,
        # Add context variables for the modal
        "income_categories": IncomeCategory.objects.all().order_by("name"),
        "accounts": BankAccount.objects.all().order_by("institution", "name"),
        "all_rental_units": RentalUnit.objects.select_related("property").order_by("property__name", "name"),
    }
    return render(request, "income_category_income_list.html", context)

# -------------------------
# Rental Properties Section
# -------------------------

def rental_properties(request):
    """
    Owned properties overview:
      - income / expenses / net over a selected year
      - mortgage summary (if configured)
      - equity, if an estimated value is set
    """
    year_param = request.GET.get("year")
    today = date.today()

    # Parse year or default to current year
    if year_param:
        try:
            selected_year = int(year_param)
        except (ValueError, TypeError):
            selected_year = today.year
    else:
        selected_year = today.year

    # Determine if YTD or full year
    is_ytd = (year_param == "ytd" or selected_year == today.year)

    if is_ytd and selected_year == today.year:
        range_start = date(today.year, 1, 1)
        range_end = today
        period_label = f"YTD {today.year}"
    else:
        range_start = date(selected_year, 1, 1)
        range_end = date(selected_year, 12, 31)
        period_label = str(selected_year)

    properties = RentalProperty.objects.filter(is_active=True).order_by("name")

    rows = []
    for prop in properties:
        income_total = (
            Income.objects.filter(
                rental_unit__property=prop,
                date__range=(range_start, range_end),
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )
        expense_total = (
            Expense.objects.filter(
                rental_unit__property=prop,
                date__range=(range_start, range_end),
            ).aggregate(total=Sum("amount"))["total"]
            or Decimal("0.00")
        )
        net_total = income_total - expense_total

        active_mortgage = prop.mortgages.filter(is_active=True).order_by("id").first()
        if active_mortgage:
            mortgage_info = {
                "name": active_mortgage.name,
                "lender_name": active_mortgage.lender_name,
                "current_balance": active_mortgage.current_principal_balance,
                "rate": active_mortgage.interest_rate_percent,
                "term_end": active_mortgage.term_end_date,
            }
        else:
            mortgage_info = None

        rows.append({
            "property": prop,
            "mortgage": mortgage_info,
            "estimated_value": prop.estimated_value,
            "equity": prop.equity,
            "ltv_pct": prop.ltv_pct,
            "income_total": income_total,
            "expense_total": expense_total,
            "net_total": net_total,
        })

    # Generate year options (current year + 5 years back)
    year_options = []
    year_options.append(("ytd", f"YTD {today.year}"))
    for y in range(today.year, today.year - 6, -1):
        year_options.append((str(y), str(y)))

    context = {
        "rows": rows,
        "selected_year": year_param if year_param else "ytd",
        "year_options": year_options,
        "range_start": range_start,
        "range_end": range_end,
        "period_label": period_label,
    }
    return render(request, "rental_properties.html", context)



def rental_property_detail(request, property_id):
    # Allow POST to carry the same month/range as GET
    selected_month_str = request.POST.get("month") or request.GET.get("month")
    range_key = request.POST.get("range") or request.GET.get("range", "6")
    today = date.today()

    try:
        year, month = map(int, (selected_month_str or "").split("-"))
    except Exception:
        year, month = today.year, today.month
        selected_month_str = f"{year:04d}-{month:02d}"

    prop = get_object_or_404(RentalProperty.objects.prefetch_related("mortgages"), pk=property_id)

    # --- Property value update (sale price / estimated value) ---
    if request.method == "POST" and request.POST.get("action") == "update_property_value":
        value_str = (request.POST.get("estimated_value") or "").strip()
        valued_date_str = (request.POST.get("last_valued_date") or "").strip()

        try:
            prop.estimated_value = Decimal(value_str) if value_str else None
        except (InvalidOperation, ValueError):
            messages.error(request, "Could not parse estimated value. Please enter a valid number.")
        else:
            if valued_date_str:
                try:
                    prop.last_valued_date = date.fromisoformat(valued_date_str)
                except ValueError:
                    messages.error(request, "Could not parse valuation date. Please use YYYY-MM-DD.")
            else:
                prop.last_valued_date = None

            prop.save()
            messages.success(request, "Property value updated.")

        # Redirect (PRG) so a refresh doesn't resubmit the form
        return redirect(f"{request.path}?month={selected_month_str}&range={range_key}")

    # -----------------------
    # Anchor month + range → list of months
    # -----------------------
    anchor_month_start = date(year, month, 1)

    def add_months(d: date, delta_months: int) -> date:
        y = d.year + (d.month - 1 + delta_months) // 12
        m = (d.month - 1 + delta_months) % 12 + 1
        return date(y, m, 1)

    month_starts = []

    if range_key in ("3", "6", "12"):
        n = int(range_key)
        start = add_months(anchor_month_start, -(n - 1))
        cur = start
        while cur <= anchor_month_start:
            month_starts.append(cur)
            cur = add_months(cur, 1)
    elif range_key == "ytd":
        start = date(anchor_month_start.year, 1, 1)
        cur = start
        while cur <= anchor_month_start:
            month_starts.append(cur)
            cur = add_months(cur, 1)
    elif range_key == "prev_year":
        prev_year = anchor_month_start.year - 1
        cur = date(prev_year, 1, 1)
        end = date(prev_year, 12, 1)
        while cur <= end:
            month_starts.append(cur)
            cur = add_months(cur, 1)
    elif range_key == "all":
        earliest_income = (
            Income.objects.filter(rental_unit__property=prop)
            .order_by("date")
            .values_list("date", flat=True)
            .first()
        )
        earliest_expense = (
            Expense.objects.filter(rental_unit__property=prop)
            .order_by("date")
            .values_list("date", flat=True)
            .first()
        )
        earliest = earliest_income or earliest_expense
        if earliest_income and earliest_expense:
            earliest = min(earliest_income, earliest_expense)
        start = date(earliest.year, earliest.month, 1) if earliest else anchor_month_start
        cur = start
        while cur <= anchor_month_start:
            month_starts.append(cur)
            cur = add_months(cur, 1)
    else:
        # default last 6 months
        start = add_months(anchor_month_start, -5)
        cur = start
        while cur <= anchor_month_start:
            month_starts.append(cur)
            cur = add_months(cur, 1)

    month_starts_chrono = sorted(month_starts)

    # -----------------------
    # Monthly cashflow + trend
    # -----------------------
    month_rows = []
    trend_labels = []
    trend_income = []
    trend_expenses = []
    trend_net = []

    for m_start in month_starts_chrono:
        start = date(m_start.year, m_start.month, 1)
        end = date(m_start.year, m_start.month, monthrange(m_start.year, m_start.month)[1])

        incomes_qs = (
            Income.objects.filter(rental_unit__property=prop, date__range=(start, end))
            .select_related("bank_account", "rental_unit", "income_category")
            .order_by("-date", "-id")
        )
        expenses_qs = (
            Expense.objects.filter(rental_unit__property=prop, date__range=(start, end))
            .select_related("bank_account", "rental_unit", "category", "cra_category")
            .order_by("-date", "-id")
        )

        income_total = incomes_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        expense_total = expenses_qs.aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
        net_total = income_total - expense_total

        month_rows.append({
            "month": m_start,
            "incomes": incomes_qs,
            "expenses": expenses_qs,
            "income_total": income_total,
            "expense_total": expense_total,
            "net_total": net_total,
        })

        trend_labels.append(m_start.strftime("%b %Y"))
        trend_income.append(float(income_total))
        trend_expenses.append(float(expense_total))
        trend_net.append(float(net_total))

    month_rows = list(reversed(month_rows))

    range_income_total = sum((r["income_total"] for r in month_rows), Decimal("0.00"))
    range_expense_total = sum((r["expense_total"] for r in month_rows), Decimal("0.00"))
    range_net_total = range_income_total - range_expense_total

    range_options = [
        ("3", "Last 3 months"),
        ("6", "Last 6 months"),
        ("12", "Last 12 months"),
        ("ytd", "Year to date"),
        ("prev_year", "Previous year"),
        ("all", "All time"),
    ]

    # -----------------------
    # Mortgage panels: YTD ledger + projections + history
    # -----------------------
    anchor_year = year
    anchor_month_end = date(year, month, monthrange(year, month)[1])

    mortgage_panels = []

    for mortgage in prop.mortgages.filter(is_active=True).order_by("name"):
        principal_categories = mortgage.principal_categories()
        interest_category_id = getattr(mortgage, "interest_category_id", None)

        principal_ytd = Decimal("0.00")
        interest_ytd = Decimal("0.00")
        ledger_rows = []
        projected_rows = []
        chart_labels = []
        chart_balances = []

        history_years = []
        history_principal = []
        history_interest = []
        history_total_principal = Decimal("0.00")
        history_pct_of_original = None

        if principal_categories:
            # --- YTD ---
            ytd_start = date(anchor_year, 1, 1)
            ytd_end = anchor_month_end

            principal_qs_ytd = Expense.objects.filter(
                category__in=principal_categories,
                date__gte=ytd_start,
                date__lte=ytd_end,
            )
            principal_ytd = principal_qs_ytd.aggregate(
                total=Coalesce(Sum("amount"), Decimal("0.00"))
            )["total"]

            if interest_category_id:
                interest_qs_ytd = Expense.objects.filter(
                    category_id=interest_category_id,
                    date__gte=ytd_start,
                    date__lte=ytd_end,
                )
                interest_ytd = interest_qs_ytd.aggregate(
                    total=Coalesce(Sum("amount"), Decimal("0.00"))
                )["total"]
            else:
                interest_qs_ytd = Expense.objects.none()
                interest_ytd = Decimal("0.00")

            # --- Ledger (current year) ---
            principal_by_date = defaultdict(lambda: Decimal("0.00"))
            prepayment_flag_by_date = defaultdict(lambda: False)
            for exp in principal_qs_ytd.select_related("category").order_by("date", "id"):
                principal_by_date[exp.date] += exp.amount
                if mortgage.prepayment_category_id and exp.category_id == mortgage.prepayment_category_id:
                    prepayment_flag_by_date[exp.date] = True

            interest_by_date = defaultdict(lambda: Decimal("0.00"))
            for exp in interest_qs_ytd.order_by("date", "id"):
                interest_by_date[exp.date] += exp.amount

            all_dates = sorted(set(principal_by_date.keys()) | set(interest_by_date.keys()))

            balance_after_by_date = {}
            base_balance = None
            base_date = None

            if (
                mortgage.tracking_start_principal is not None
                and mortgage.tracking_start_date is not None
                and mortgage.tracking_start_date in all_dates
            ):
                base_date = mortgage.tracking_start_date
                base_balance = (
                    mortgage.tracking_start_principal
                    + (mortgage.manual_adjustment or Decimal("0.00"))
                )
            elif mortgage.tracking_start_principal is not None and mortgage.tracking_start_date is not None and all_dates:
                base_date = all_dates[0]
                base_balance = (
                    mortgage.tracking_start_principal
                    + (mortgage.manual_adjustment or Decimal("0.00"))
                )

            if base_balance is not None and all_dates:
                if base_date in all_dates:
                    idx = all_dates.index(base_date)
                    balance_after_by_date[base_date] = base_balance

                    # Forward (later dates): subtract that day's principal
                    for i in range(idx + 1, len(all_dates)):
                        d = all_dates[i]
                        prev_d = all_dates[i - 1]
                        prev_bal = balance_after_by_date[prev_d]
                        balance_after_by_date[d] = prev_bal - principal_by_date[d]

                    # Backward (earlier dates): add next day's principal
                    for i in range(idx - 1, -1, -1):
                        d = all_dates[i]
                        next_d = all_dates[i + 1]
                        next_bal = balance_after_by_date[next_d]
                        balance_after_by_date[d] = next_bal + principal_by_date[next_d]
                else:
                    # Simple forward pass (no good anchor)
                    bal = base_balance
                    for d in all_dates:
                        bal = bal - principal_by_date[d]
                        balance_after_by_date[d] = bal

            for d in all_dates:
                principal_amt = principal_by_date[d]
                interest_amt = interest_by_date[d]
                payment_total = principal_amt + interest_amt

                if prepayment_flag_by_date[d] and interest_amt == 0:
                    payment_type = "Prepayment"
                elif principal_amt > 0 and interest_amt > 0:
                    payment_type = "Regular payment"
                elif principal_amt > 0:
                    payment_type = "Principal-only"
                elif interest_amt > 0:
                    payment_type = "Interest-only"
                else:
                    payment_type = "Other"

                balance_after = balance_after_by_date.get(d)

                ledger_rows.append({
                    "date": d,
                    "payment_type": payment_type,
                    "principal": principal_amt,
                    "interest": interest_amt,
                    "payment_total": payment_total,
                    "balance_after": balance_after,
                })

            # --- Projections ---
            last_balance = None
            last_payment_date = None
            for row in ledger_rows:
                if row["balance_after"] is not None:
                    last_balance = row["balance_after"]
                    last_payment_date = row["date"]

            if last_balance is not None and last_payment_date is not None:
                projected_rows = mortgage.projected_rows_to_year_end(
                    starting_balance=last_balance,
                    last_payment_date=last_payment_date,
                    year=anchor_year,
                )

            # --- Chart data (current year + projections) ---
            for row in ledger_rows:
                if row["balance_after"] is not None:
                    chart_labels.append(row["date"].isoformat())
                    chart_balances.append(float(row["balance_after"]))
            for row in projected_rows:
                chart_labels.append(row["date"].isoformat())
                chart_balances.append(float(row["balance_after"]))

            # --- History: principal & interest by year (all time) ---
            principal_all = Expense.objects.filter(category__in=principal_categories)
            principal_by_year = defaultdict(lambda: Decimal("0.00"))
            for exp in principal_all.only("date", "amount"):
                principal_by_year[exp.date.year] += exp.amount

            interest_by_year = defaultdict(lambda: Decimal("0.00"))
            if interest_category_id:
                interest_all = Expense.objects.filter(category_id=interest_category_id)
                for exp in interest_all.only("date", "amount"):
                    interest_by_year[exp.date.year] += exp.amount

            all_years = sorted(set(principal_by_year.keys()) | set(interest_by_year.keys()))
            running = Decimal("0.00")
            for y in all_years:
                p = principal_by_year[y]
                i = interest_by_year[y]
                running += p
                history_years.append(y)
                history_principal.append(float(p))
                history_interest.append(float(i))
                history_total_principal = running

            if mortgage.original_principal and mortgage.original_principal > 0 and history_total_principal > 0:
                history_pct_of_original = (history_total_principal / mortgage.original_principal) * Decimal("100.0")

        mortgage_panels.append({
            "mortgage": mortgage,
            "principal_ytd": principal_ytd,
            "interest_ytd": interest_ytd,
            "ledger_rows": ledger_rows,
            "projected_rows": projected_rows,
            "chart_labels": chart_labels,
            "chart_balances": chart_balances,
            "history_years": history_years,
            "history_principal": history_principal,
            "history_interest": history_interest,
            "history_total_principal": history_total_principal,
            "history_pct_of_original": history_pct_of_original,
        })

    context = {
        "property": prop,
        "selected_month": f"{year:04d}-{month:02d}",
        "selected_range": range_key,
        "range_options": range_options,
        "month_rows": month_rows,
        "trend_labels": trend_labels,
        "trend_income": trend_income,
        "trend_expenses": trend_expenses,
        "trend_net": trend_net,
        "range_income_total": range_income_total,
        "range_expense_total": range_expense_total,
        "range_net_total": range_net_total,
        "mortgage_panels": mortgage_panels,
    }
    return render(request, "rental_property_detail.html", context)







def rental_tax_summary(request, property_id):
    """
    Tax Summary (per property, per year) in a CRA-ish layout:
      - Income: gross rents (all units), unit count (excludes Shared/Common)
      - Expenses: list ALL CRA categories (even if $0)
      - For each CRA category:
          Total expenses (raw)
          Personal portion (raw - rental portion)
          Rental portion (after rental_business_use_pct; default 100%)
      - Drilldown link per CRA category
    """
    prop = get_object_or_404(RentalProperty, pk=property_id)

    # ---- Year dropdown options (earliest -> current) ----
    current_year = date.today().year

    earliest_expense = (
        Expense.objects
        .filter(rental_unit__property=prop)
        .order_by("date")
        .values_list("date", flat=True)
        .first()
    )
    earliest_income = (
        Income.objects
        .filter(rental_unit__property=prop)
        .order_by("date")
        .values_list("date", flat=True)
        .first()
    )

    earliest = earliest_expense or earliest_income
    if earliest_expense and earliest_income:
        earliest = min(earliest_expense, earliest_income)

    start_year = earliest.year if earliest else current_year
    year_options = list(range(start_year, current_year + 1))

    try:
        year = int(request.GET.get("year") or current_year)
    except ValueError:
        year = current_year

    if year not in year_options:
        year_options.append(year)
        year_options = sorted(set(year_options))

    # ---- Unit count (exclude Shared/Common units) ----
    units_qs = RentalUnit.objects.filter(property=prop, is_active=True)
    # simple "smart enough" heuristic
    rentable_units = units_qs.exclude(name__icontains="shared").exclude(name__icontains="common")
    unit_count = rentable_units.count()

    # ---- Income: gross rents for all units (we ignore short-term rentals) ----
    # We assume rental incomes are already tagged with rental_unit.
    income_total = (
        Income.objects
        .filter(rental_unit__property=prop, date__year=year)
        .aggregate(total=Sum("amount"))["total"]
        or Decimal("0.00")
    )

    # ---- Expenses: only those with CRA category set (for summary table) ----
    expense_qs = (
        Expense.objects
        .filter(
            date__year=year,
            rental_unit__property=prop,
            cra_category__isnull=False,
        )
        .select_related("cra_category", "rental_unit")
    )

    # Rental %: default 100 if null
    pct = Case(
        When(rental_business_use_pct__isnull=False, then=F("rental_business_use_pct")),
        default=Value(100),
        output_field=DecimalField(max_digits=5, decimal_places=2),
    )

    rental_portion_expr = ExpressionWrapper(
        F("amount") * pct / Value(100),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )

    personal_portion_expr = ExpressionWrapper(
        F("amount") - (F("amount") * pct / Value(100)),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )

    # Aggregate by CRA category id so we can join to the full category list (including zeros)
    aggregates = (
        expense_qs.values("cra_category_id")
        .annotate(
            total_raw=Sum("amount"),
            total_rental=Sum(rental_portion_expr),
            total_personal=Sum(personal_portion_expr),
        )
    )

    agg_map = {
        row["cra_category_id"]: row
        for row in aggregates
    }

    # ---- Build rows for ALL CRA categories (even if zero) ----
    cra_categories = list(
        CRARentalExpenseCategory.objects
        .filter(is_active=True)
        .order_by("sort_order", "name")
    )

    rows = []
    total_expenses_raw = Decimal("0.00")
    total_expenses_rental = Decimal("0.00")
    total_expenses_personal = Decimal("0.00")

    for cat in cra_categories:
        data = agg_map.get(cat.id) or {}
        raw = data.get("total_raw") or Decimal("0.00")
        rental = data.get("total_rental") or Decimal("0.00")
        personal = data.get("total_personal") or Decimal("0.00")

        total_expenses_raw += raw
        total_expenses_rental += rental
        total_expenses_personal += personal

        rows.append({
            "cat": cat,
            "total_raw": raw,
            "total_rental": rental,
            "total_personal": personal,
        })

    # How many expenses are missing CRA category (for cleanup)
    missing_cra_count = (
        Expense.objects
        .filter(date__year=year, rental_unit__property=prop, cra_category__isnull=True)
        .count()
    )
    missing_cra_expenses = (
        Expense.objects
        .filter(
            rental_unit__property=prop,
            date__year=year,
            cra_category__isnull=True,
        )
        .select_related("category", "rental_unit", "bank_account")
        .order_by("-date", "-id")
    )
    missing_cra_count = missing_cra_expenses.count()

    # ---- Expenses with CRA category but NO receipts (important for tax audit) ----
    expenses_missing_receipts = (
        Expense.objects
        .filter(
            rental_unit__property=prop,
            date__year=year,
            cra_category__isnull=False,  # Has CRA category (tax-deductible)
        )
        .annotate(attachment_count=Count("attachments"))
        .filter(attachment_count=0)  # But no receipts
        .select_related("category", "rental_unit", "bank_account", "cra_category")
        .order_by("-date", "-id")
    )
    missing_receipts_count = expenses_missing_receipts.count()

    context = {
        "property": prop,
        "year": year,
        "year_options": year_options,
        "unit_count": unit_count,
        "income_total": income_total,
        "rows": rows,
        "total_expenses_raw": total_expenses_raw,
        "total_expenses_rental": total_expenses_rental,
        "total_expenses_personal": total_expenses_personal,
        "missing_cra_count": missing_cra_count,
        "missing_cra_expenses": missing_cra_expenses,
        "expenses_missing_receipts": expenses_missing_receipts,
        "missing_receipts_count": missing_receipts_count,
    }
    return render(request, "rental_tax_summary.html", context)


def rental_tax_export(request, property_id):
    """
    Export a ZIP containing:
      - An Excel workbook with the CRA Part 3/4 summary + per-transaction breakdown
      - All expense receipts organized by CRA category folder
    """
    import zipfile
    import os
    import re
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    prop = get_object_or_404(RentalProperty, pk=property_id)

    try:
        year = int(request.GET.get("year") or date.today().year)
    except ValueError:
        year = date.today().year

    # ---- Shared ORM expressions (same as tax_summary view) ----
    pct = Case(
        When(rental_business_use_pct__isnull=False, then=F("rental_business_use_pct")),
        default=Value(100),
        output_field=DecimalField(max_digits=5, decimal_places=2),
    )
    rental_portion_expr = ExpressionWrapper(
        F("amount") * pct / Value(100),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    personal_portion_expr = ExpressionWrapper(
        F("amount") - (F("amount") * pct / Value(100)),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )

    # Income
    income_total = (
        Income.objects
        .filter(rental_unit__property=prop, date__year=year)
        .aggregate(total=Sum("amount"))["total"] or Decimal("0.00")
    )

    # Unit count
    unit_count = (
        RentalUnit.objects
        .filter(property=prop, is_active=True)
        .exclude(name__icontains="shared")
        .exclude(name__icontains="common")
        .count()
    )

    # Expenses with CRA category
    expense_qs = (
        Expense.objects
        .filter(date__year=year, rental_unit__property=prop, cra_category__isnull=False)
        .select_related("cra_category", "rental_unit", "category")
        .prefetch_related("attachments")
        .order_by("cra_category__sort_order", "cra_category__name", "date")
    )

    # Aggregate totals by CRA category
    # Must call .order_by() first to clear the queryset ordering — otherwise Django
    # includes the order_by fields (e.g. "date") in GROUP BY, splitting rows incorrectly.
    aggregates = (
        expense_qs.order_by()
        .values("cra_category_id")
        .annotate(
            total_raw=Sum("amount"),
            total_rental=Sum(rental_portion_expr),
            total_personal=Sum(personal_portion_expr),
        )
    )
    agg_map = {row["cra_category_id"]: row for row in aggregates}

    cra_categories = list(
        CRARentalExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name")
    )

    rows = []
    total_raw = Decimal("0.00")
    total_rental = Decimal("0.00")
    total_personal = Decimal("0.00")
    for cat in cra_categories:
        data = agg_map.get(cat.id) or {}
        r = data.get("total_raw") or Decimal("0.00")
        rn = data.get("total_rental") or Decimal("0.00")
        rp = data.get("total_personal") or Decimal("0.00")
        total_raw += r
        total_rental += rn
        total_personal += rp
        rows.append({"cat": cat, "total_raw": r, "total_rental": rn, "total_personal": rp})

    # Group individual expenses by CRA category id
    expenses_by_cat = defaultdict(list)
    for exp in expense_qs:
        expenses_by_cat[exp.cra_category_id].append(exp)

    # ---- Income detail rows ----
    income_entries = (
        Income.objects
        .filter(rental_unit__property=prop, date__year=year)
        .select_related("income_category", "rental_unit")
        .order_by("date")
    )

    # ---- Mortgage data: principal + interest by month ----
    import calendar as cal_mod
    mortgages = list(
        PropertyMortgage.objects
        .filter(owned_property=prop)
        .select_related("principal_category", "prepayment_category", "interest_category")
    )
    MONTHS = [cal_mod.month_name[m] for m in range(1, 13)]

    mortgage_sections = []
    for mortgage in mortgages:
        principal_cats = mortgage.principal_categories()  # list of Category objects
        interest_cat = mortgage.interest_category

        monthly = []
        total_principal = Decimal("0.00")
        total_interest = Decimal("0.00")

        for month_num in range(1, 13):
            p = Decimal("0.00")
            i = Decimal("0.00")
            if principal_cats:
                p = (
                    Expense.objects
                    .filter(category__in=principal_cats, date__year=year, date__month=month_num)
                    .aggregate(t=Coalesce(Sum("amount"), Decimal("0.00")))["t"]
                )
            if interest_cat:
                i = (
                    Expense.objects
                    .filter(category=interest_cat, date__year=year, date__month=month_num)
                    .aggregate(t=Coalesce(Sum("amount"), Decimal("0.00")))["t"]
                )
            total_principal += p
            total_interest += i
            monthly.append({"month": MONTHS[month_num - 1], "principal": p, "interest": i, "total": p + i})

        mortgage_sections.append({
            "mortgage": mortgage,
            "monthly": monthly,
            "total_principal": total_principal,
            "total_interest": total_interest,
            "total": total_principal + total_interest,
        })

    generated_at = datetime.now()
    generated_str = generated_at.strftime("%Y-%m-%d %H:%M")
    generated_slug = generated_at.strftime("%Y%m%d_%H%M%S")

    # ------------------------------------------------------------------ #
    #  BUILD EXCEL                                                         #
    # ------------------------------------------------------------------ #
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Tax Summary {year}"

    # Styles
    def h1(text):
        return Font(bold=True, size=14)

    def h2():
        return Font(bold=True, size=11, color="FFFFFF")

    green_fill = PatternFill("solid", fgColor="1F7A4F")
    red_fill   = PatternFill("solid", fgColor="C0392B")
    gray_fill  = PatternFill("solid", fgColor="D9D9D9")
    sub_fill   = PatternFill("solid", fgColor="F2F2F2")
    total_fill = PatternFill("solid", fgColor="BDD7EE")
    thin_side  = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    col_widths = [40, 18, 25, 25, 20, 20, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    def write_row(values, fill=None, bold=False, row_num=None):
        r = row_num or row
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=False, vertical="center")
            if fill:
                cell.fill = fill
            if bold:
                cell.font = Font(bold=True)
            if isinstance(val, Decimal):
                cell.value = float(val)
                cell.number_format = '"$"#,##0.00'
        return r

    # Title
    ws.merge_cells(f"A{row}:G{row}")
    title_cell = ws.cell(row=row, column=1, value=f"Rental Tax Summary — {prop.name} — {year}")
    title_cell.font = Font(bold=True, size=15)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 24
    row += 1

    ws.merge_cells(f"A{row}:G{row}")
    gen_cell = ws.cell(row=row, column=1, value=f"Generated: {generated_str}")
    gen_cell.font = Font(italic=True, size=10, color="888888")
    gen_cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 2

    blue_fill   = PatternFill("solid", fgColor="1A5276")

    # ---- Part 3: Income ----
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PART 3 — INCOME")
    c.font = h2()
    c.fill = green_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    write_row(["Property", prop.name, "", "", "", "", ""], fill=sub_fill, row_num=row); row += 1
    write_row(["Number of rentable units", unit_count, "", "", "", "", ""], fill=sub_fill, row_num=row); row += 1
    write_row(["Gross rental income", income_total, "", "", "", "", ""], fill=sub_fill, bold=True, row_num=row); row += 1
    row += 1

    # Income transaction detail
    inc_headers = ["Date", "Income Category", "Rental Unit", "Amount", "Taxable", "Notes", ""]
    for col, h in enumerate(inc_headers[:6], 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    income_year_total = Decimal("0.00")
    for entry in income_entries:
        cat_name = entry.income_category.name if entry.income_category else (entry.category or "—")
        unit_name = entry.rental_unit.name if entry.rental_unit else "—"
        income_year_total += entry.amount
        values = [
            entry.date.strftime("%Y-%m-%d"),
            cat_name,
            unit_name,
            float(entry.amount),
            "Yes" if entry.taxable else "No",
            entry.notes or "",
            "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            if col == 4:
                cell.number_format = '"$"#,##0.00'
        row += 1

    # Income total row
    for col, val in [(1, "Total Income"), (4, float(income_year_total))]:
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin_border
        if col == 4:
            cell.number_format = '"$"#,##0.00'
    row += 2

    # ---- Mortgage Payments ----
    if mortgage_sections:
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value="MORTGAGE PAYMENTS")
        c.font = h2()
        c.fill = blue_fill
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for ms in mortgage_sections:
            m = ms["mortgage"]
            label = m.name
            if m.lender_name:
                label = f"{m.lender_name} — {label}"

            # Mortgage name sub-header
            ws.merge_cells(f"A{row}:G{row}")
            c = ws.cell(row=row, column=1, value=label)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="555555")
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

            mort_headers = ["Month", "Principal", "Interest", "Total Payment", "", "", ""]
            for col, h in enumerate(mort_headers[:4], 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = Font(bold=True)
                cell.fill = gray_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

            for mrow in ms["monthly"]:
                # Skip months with no activity
                if mrow["total"] == Decimal("0.00"):
                    continue
                for col, val in [(1, mrow["month"]), (2, float(mrow["principal"])),
                                 (3, float(mrow["interest"])), (4, float(mrow["total"]))]:
                    cell = ws.cell(row=row, column=col, value=val)
                    cell.border = thin_border
                    if col > 1:
                        cell.number_format = '"$"#,##0.00'
                row += 1

            # Annual totals
            for col, val in [(1, f"Annual Total — {m.name}"),
                             (2, float(ms["total_principal"])),
                             (3, float(ms["total_interest"])),
                             (4, float(ms["total"]))]:
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = thin_border
                if col > 1:
                    cell.number_format = '"$"#,##0.00'
            row += 2

    # ---- Part 4: Summary table ----
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PART 4 — EXPENSES SUMMARY")
    c.font = h2()
    c.fill = red_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    # Header row
    headers = ["CRA Expense Category", "Total Expenses", "Personal Portion", "Rental Portion", "", "", ""]
    for col, val in enumerate(headers[:4], 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True)
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    for r in rows:
        ws.cell(row=row, column=1, value=r["cat"].name).border = thin_border
        for col, val in [(2, r["total_raw"]), (3, r["total_personal"]), (4, r["total_rental"])]:
            cell = ws.cell(row=row, column=col, value=float(val))
            cell.number_format = '"$"#,##0.00'
            cell.border = thin_border
        row += 1

    # Totals row
    for col, val in [(1, "TOTAL"), (2, total_raw), (3, total_personal), (4, total_rental)]:
        cell = ws.cell(row=row, column=col, value=val if col == 1 else float(val))
        cell.font = Font(bold=True)
        cell.fill = total_fill
        cell.border = thin_border
        if col > 1:
            cell.number_format = '"$"#,##0.00'
    row += 2

    # ---- Per-category transaction breakdown ----
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value="PART 4 — TRANSACTION DETAIL BY CATEGORY")
    c.font = h2()
    c.fill = red_fill
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    txn_headers = ["Date", "Vendor", "Category (internal)", "Rental Unit", "Amount", "Rental Portion", "Notes"]
    for cat_row in rows:
        cat = cat_row["cat"]
        expenses = expenses_by_cat.get(cat.id, [])
        if not expenses:
            continue

        # Category sub-header
        ws.merge_cells(f"A{row}:G{row}")
        c = ws.cell(row=row, column=1, value=cat.name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="555555")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # Column headers
        for col, h in enumerate(txn_headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = gray_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
        row += 1

        cat_total_raw = Decimal("0.00")
        cat_total_rental = Decimal("0.00")
        for exp in expenses:
            pct_val = exp.rental_business_use_pct if exp.rental_business_use_pct is not None else Decimal("100")
            rental_amt = (exp.amount * pct_val / 100).quantize(Decimal("0.01"))
            cat_total_raw += exp.amount
            cat_total_rental += rental_amt

            values = [
                exp.date.strftime("%Y-%m-%d"),
                exp.vendor_name,
                exp.category.name if exp.category else "",
                exp.rental_unit.name if exp.rental_unit else "",
                float(exp.amount),
                float(rental_amt),
                exp.notes or "",
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
                if col in (5, 6):
                    cell.number_format = '"$"#,##0.00'
            row += 1

        # Category subtotal
        for col, val in [(1, f"Subtotal — {cat.name}"), (5, float(cat_total_raw)), (6, float(cat_total_rental))]:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(bold=True)
            cell.fill = sub_fill
            cell.border = thin_border
            if col in (5, 6):
                cell.number_format = '"$"#,##0.00'
        row += 2

    # Save workbook to bytes
    xlsx_buf = BytesIO()
    wb.save(xlsx_buf)
    xlsx_bytes = xlsx_buf.getvalue()

    # ------------------------------------------------------------------ #
    #  BUILD ZIP                                                           #
    # ------------------------------------------------------------------ #
    def safe_folder(name):
        return re.sub(r'[\\/:*?"<>|]', '_', name).strip()

    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        # Excel file
        safe_prop = safe_folder(prop.name)
        zf.writestr(f"Tax_Summary_{safe_prop}_{year}_{generated_slug}.xlsx", xlsx_bytes)

        # Receipts organized by CRA category
        for cat_row in rows:
            cat = cat_row["cat"]
            expenses = expenses_by_cat.get(cat.id, [])
            for exp in expenses:
                for attachment in exp.attachments.all():
                    file_path = os.path.join(settings.MEDIA_ROOT, attachment.file.name)
                    if not os.path.exists(file_path):
                        continue
                    _, ext = os.path.splitext(attachment.file.name)
                    receipt_name = (
                        f"{exp.date.strftime('%Y-%m-%d')}_"
                        f"{safe_folder(exp.vendor_name)}_"
                        f"${exp.amount:.2f}{ext}"
                    )
                    zip_path = f"receipts/{safe_folder(cat.name)}/{receipt_name}"
                    with open(file_path, "rb") as fh:
                        zf.writestr(zip_path, fh.read())

    zip_bytes = zip_buf.getvalue()

    response = HttpResponse(zip_bytes, content_type="application/zip")
    response["Content-Disposition"] = (
        f'attachment; filename="TaxExport_{safe_folder(prop.name)}_{year}_{generated_slug}.zip"'
    )
    return response


def rental_tax_category_detail(request, property_id, cra_category_id):
    prop = get_object_or_404(RentalProperty, pk=property_id)
    cra_cat = get_object_or_404(CRARentalExpenseCategory, pk=cra_category_id)

    try:
        year = int(request.GET.get("year") or date.today().year)
    except ValueError:
        year = date.today().year

    # POST: Handle edit/delete
    if request.method == "POST":
        expense_id = request.POST.get("expense_id")
        if expense_id:
            exp_obj = get_object_or_404(Expense, pk=expense_id)

            # Delete if requested
            if request.POST.get("delete_expense"):
                exp_obj.delete()
                return redirect(f"{reverse('rental_tax_category_detail', args=[property_id, cra_category_id])}?year={year}")

            # Otherwise, update
            exp_obj.date = request.POST.get("date") or exp_obj.date
            exp_obj.vendor_name = request.POST.get("vendor_name") or ""
            exp_obj.location = request.POST.get("location") or ""
            exp_obj.notes = request.POST.get("notes") or ""

            try:
                exp_obj.amount = Decimal(request.POST.get("amount", "0"))
            except (ValueError, InvalidOperation):
                pass

            # Category
            cat_name = request.POST.get("category", "").strip()
            if cat_name:
                cat_obj, _ = Category.objects.get_or_create(name=cat_name)
                exp_obj.category = cat_obj

            # Bank account
            ba_id = request.POST.get("bank_account")
            if ba_id:
                exp_obj.bank_account = BankAccount.objects.get(pk=ba_id)
            else:
                exp_obj.bank_account = None

            # Rental unit
            ru_id = request.POST.get("rental_unit")
            if ru_id:
                exp_obj.rental_unit = RentalUnit.objects.get(pk=ru_id)
            else:
                exp_obj.rental_unit = None

            # CRA category
            cra_id = request.POST.get("cra_category")
            if cra_id:
                exp_obj.cra_category = CRARentalExpenseCategory.objects.get(pk=cra_id)
            else:
                exp_obj.cra_category = None

            exp_obj.save()
            return redirect(f"{reverse('rental_tax_category_detail', args=[property_id, cra_category_id])}?year={year}")

    # GET: Display expenses
    qs = (
        Expense.objects
        .filter(
            rental_unit__property=prop,
            date__year=year,
            cra_category=cra_cat,
        )
        .select_related("rental_unit", "category", "bank_account")
        .prefetch_related("attachments")
        .order_by("-date", "-id")
    )

    pct = Case(
        When(rental_business_use_pct__isnull=False, then=F("rental_business_use_pct")),
        default=Value(100),
        output_field=DecimalField(max_digits=5, decimal_places=2),
    )

    rental_portion_expr = ExpressionWrapper(
        F("amount") * pct / Value(100),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )
    personal_portion_expr = ExpressionWrapper(
        F("amount") - (F("amount") * pct / Value(100)),
        output_field=DecimalField(max_digits=12, decimal_places=2),
    )

    qs = qs.annotate(
        rental_portion=rental_portion_expr,
        personal_portion=personal_portion_expr,
    )

    # Context for modal form
    all_categories = Category.objects.filter(is_archived=False).order_by("name")
    accounts = BankAccount.objects.all().order_by("name")
    all_rental_units = RentalUnit.objects.select_related("property").all().order_by("property__name", "name")
    cra_categories = CRARentalExpenseCategory.objects.filter(is_active=True).order_by("sort_order", "name")

    context = {
        "property": prop,
        "cra_category": cra_cat,
        "year": year,
        "expenses": qs,
        "all_categories": all_categories,
        "accounts": accounts,
        "all_rental_units": all_rental_units,
        "cra_categories": cra_categories,
    }
    return render(request, "rental_tax_category_detail.html", context)


# -------------------------
# Categories + Accounts (unchanged)
# -------------------------

class CategoryForm(ModelForm):
    class Meta:
        model = Category
        fields = ["name", "monthly_limit"]

class CategoryForm(ModelForm):
    class Meta:
        model = Category
        fields = ["name", "monthly_limit"]

class IncomeCategoryForm(ModelForm):
    class Meta:
        model = IncomeCategory
        fields = ["name", "monthly_target", "taxable_default"]

class WithholdingCategoryForm(ModelForm):
    class Meta:
        model = WithholdingCategory
        fields = ["name", "account", "monthly_target", "target_amount", "next_due_date"]
        widgets = {
            "next_due_date": forms.DateInput(attrs={"type": "date"}),
        }

def category_list(request):
    expense_categories = Category.objects.filter(is_archived=False).order_by("name")
    archived_categories = Category.objects.filter(is_archived=True).order_by("name")
    income_categories = IncomeCategory.objects.all().order_by("name")
    withholding_categories = WithholdingCategory.objects.select_related("account").all().order_by("name")

    if request.method == "POST":
        action = request.POST.get("action")

        if action == "expense_archive":
            cat = get_object_or_404(Category, pk=request.POST.get("id"))
            cat.is_archived = True
            cat.save()
            return redirect("category_list")

        if action == "expense_unarchive":
            cat = get_object_or_404(Category, pk=request.POST.get("id"))
            cat.is_archived = False
            cat.save()
            return redirect("category_list")

        if action == "expense_delete":
            cat = get_object_or_404(Category, pk=request.POST.get("id"))
            cat.delete()
            return redirect("category_list")

        if action == "expense_save":
            category_id = request.POST.get("id") or ""
            instance = get_object_or_404(Category, pk=category_id) if category_id else None
            form = CategoryForm(request.POST, instance=instance)

            if form.is_valid():
                try:
                    form.save()
                    return redirect("category_list")
                except IntegrityError:
                    messages.error(request, "An expense category with that name already exists.")
            else:
                messages.error(request, "Please correct the errors in the expense category form.")

        if action == "income_delete":
            inc = get_object_or_404(IncomeCategory, pk=request.POST.get("id"))
            inc.delete()
            return redirect("category_list")

        if action == "income_save":
            income_id = request.POST.get("id") or ""
            instance = get_object_or_404(IncomeCategory, pk=income_id) if income_id else None
            form = IncomeCategoryForm(request.POST, instance=instance)

            if form.is_valid():
                try:
                    form.save()
                    return redirect("category_list")
                except IntegrityError:
                    messages.error(request, "An income category with that name already exists.")
            else:
                messages.error(request, "Please correct the errors in the income category form.")

        if action == "withholding_delete":
            wh = get_object_or_404(WithholdingCategory, pk=request.POST.get("id"))
            wh.delete()
            return redirect("category_list")

        if action == "withholding_save":
            withholding_id = request.POST.get("id") or ""
            instance = get_object_or_404(WithholdingCategory, pk=withholding_id) if withholding_id else None
            form = WithholdingCategoryForm(request.POST, instance=instance)

            if form.is_valid():
                try:
                    form.save()
                    return redirect("category_list")
                except IntegrityError:
                    messages.error(request, "A withholding category with that name already exists.")
            else:
                messages.error(request, "Please correct the errors in the withholding category form.")

    expense_form = CategoryForm()
    income_form = IncomeCategoryForm()
    withholding_form = WithholdingCategoryForm()

    return render(request, "category_list.html", {
        "expense_categories": expense_categories,
        "archived_categories": archived_categories,
        "income_categories": income_categories,
        "withholding_categories": withholding_categories,
        "expense_form": expense_form,
        "income_form": income_form,
        "withholding_form": withholding_form,
    })

class BankAccountForm(ModelForm):
    class Meta:
        model = BankAccount
        fields = [
            "name",
            "institution",
            "account_type",
            "account_number_last4",
            "is_withholding_account",
            "current_balance",
            "is_active",
        ]

def bank_accounts(request):
    accounts = BankAccount.objects.all().order_by("name")

    if request.method == "POST":
        if "delete_account" in request.POST:
            account = get_object_or_404(BankAccount, pk=request.POST.get("account_id"))
            account.delete()
            return redirect("bank_accounts")

        account_id = request.POST.get("account_id")
        if account_id:
            account = get_object_or_404(BankAccount, pk=account_id)
            form = BankAccountForm(request.POST, instance=account)
        else:
            form = BankAccountForm(request.POST)

        if form.is_valid():
            account = form.save(commit=False)
            account.last_updated = date.today()
            account.save()
            return redirect("bank_accounts")
    else:
        form = BankAccountForm()

    return render(request, "bank_accounts.html", {"accounts": accounts, "form": form})

def bank_account_detail(request, account_id):
    """
    Per-account ledger view.

    Shows transactions for this account with running balance.
    Defaults to last 12 months, grouped by month.
    """
    account = get_object_or_404(BankAccount, pk=account_id)

    # --- Date range selection ---
    today = date.today()

    # Default to last 12 months
    default_start = date(today.year - 1, today.month, 1)
    default_end = date(today.year, today.month, monthrange(today.year, today.month)[1])

    # Allow override via query params
    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    if start_date_str and end_date_str:
        try:
            first_day = datetime.strptime(start_date_str, "%Y-%m-%d").date()
            last_day = datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            first_day = default_start
            last_day = default_end
    else:
        first_day = default_start
        last_day = default_end

    date_range_display = f"{first_day.strftime('%b %Y')} - {last_day.strftime('%b %Y')}"

    # --- Pull all transactions for this account in date range ---
    income_qs = (
        account.incomes.filter(date__range=(first_day, last_day))
        .select_related("income_category", "rental_unit")
        .order_by("date", "id")
    )

    expense_qs = (
        account.expenses.filter(date__range=(first_day, last_day))
        .select_related("category", "rental_unit")
        .order_by("date", "id")
    )

    incoming_transfers = (
        account.incoming_transfers.filter(date__range=(first_day, last_day))
        .select_related("from_account", "to_account", "withholding_category")
        .order_by("date", "id")
    )

    outgoing_transfers = (
        account.outgoing_transfers.filter(date__range=(first_day, last_day))
        .select_related("from_account", "to_account", "withholding_category")
        .order_by("date", "id")
    )

    # Balance adjustments
    balance_adjustments = (
        account.balance_adjustments.filter(date__range=(first_day, last_day))
        .order_by("date", "id")
    )

    # --- Normalize to a single list of ledger entries ---
    entries = []

    # Incomes (inflows)
    for inc in income_qs:
        desc = ""
        if inc.income_category:
            desc = inc.income_category.name
        elif inc.category:
            desc = inc.category
        else:
            desc = "Income"

        if inc.rental_unit:
            desc = f"{desc} – {inc.rental_unit.property.name} / {inc.rental_unit.name}"

        entries.append(
            {
                "date": inc.date,
                "kind": "income",
                "is_inflow": True,
                "raw_amount": inc.amount,
                "signed_amount": inc.amount,  # inflow = +amount
                "description": desc,
                "notes": inc.notes,
                "income_id": inc.id,
                "expense_id": None,
                "transfer_id": None,
                "adjustment_id": None,
            }
        )

    # Expenses (outflows)
    for exp in expense_qs:
        desc = exp.vendor_name or ""
        if exp.category:
            if desc:
                desc = f"{desc} – {exp.category.name}"
            else:
                desc = exp.category.name

        if exp.rental_unit:
            tail = f"{exp.rental_unit.property.name} / {exp.rental_unit.name}"
            desc = f"{desc} ({tail})" if desc else tail

        entries.append(
            {
                "date": exp.date,
                "kind": "expense",
                "is_inflow": False,
                "raw_amount": exp.amount,
                "signed_amount": -exp.amount,  # outflow = -amount
                "description": desc or "Expense",
                "notes": exp.notes,
                "income_id": None,
                "expense_id": exp.id,
                "transfer_id": None,
                "adjustment_id": None,
            }
        )

    # Transfers where this account is the *destination* (inflows)
    for tr in incoming_transfers:
        # If someone ever recorded from_account == to_account == this account, skip (no net effect)
        if tr.from_account_id == account.id and tr.to_account_id == account.id:
            continue

        counterparty = tr.from_account or None
        base_desc = tr.description or "Transfer in"

        if counterparty:
            desc = f"{base_desc} (from {counterparty})"
        else:
            desc = f"{base_desc} (from external)"

        if tr.withholding_category:
            desc = f"{desc} [Bucket: {tr.withholding_category.name}]"

        entries.append(
            {
                "date": tr.date,
                "kind": "transfer",
                "is_inflow": True,
                "raw_amount": tr.amount,
                "signed_amount": tr.amount,  # inflow = +amount
                "description": desc,
                "notes": tr.notes,
                "income_id": None,
                "expense_id": None,
                "transfer_id": tr.id,
                "adjustment_id": None,
            }
        )

    # Transfers where this account is the *source* (outflows)
    for tr in outgoing_transfers:
        # If someone ever recorded from_account == to_account == this account, skip (no net effect)
        if tr.from_account_id == account.id and tr.to_account_id == account.id:
            continue

        counterparty = tr.to_account or None
        base_desc = tr.description or "Transfer out"

        if counterparty:
            desc = f"{base_desc} (to {counterparty})"
        else:
            desc = f"{base_desc} (to external)"

        if tr.withholding_category:
            desc = f"{desc} [Bucket: {tr.withholding_category.name}]"

        entries.append(
            {
                "date": tr.date,
                "kind": "transfer",
                "is_inflow": False,
                "raw_amount": tr.amount,
                "signed_amount": -tr.amount,  # outflow = -amount
                "description": desc,
                "notes": tr.notes,
                "income_id": None,
                "expense_id": None,
                "transfer_id": tr.id,
                "adjustment_id": None,
            }
        )

    # Balance adjustments
    for adj in balance_adjustments:
        entries.append(
            {
                "date": adj.date,
                "kind": "adjustment",
                "is_inflow": adj.amount >= 0,
                "raw_amount": abs(adj.amount),
                "signed_amount": adj.amount,
                "description": f"Balance Adjustment: {adj.reason}",
                "notes": adj.notes,
                "income_id": None,
                "expense_id": None,
                "transfer_id": None,
                "adjustment_id": adj.id,
            }
        )

    # Sort chronologically (oldest first for balance calculation)
    def sort_key(e):
        kind_order = {"income": 0, "transfer": 1, "expense": 2, "adjustment": 3}
        return (e["date"], kind_order.get(e["kind"], 99))

    entries.sort(key=sort_key)  # Oldest first for running balance

    # --- Calculate running balance ---
    # For balance tracking: use February 2026 snapshot as the baseline
    if account.balance_tracking_enabled and account.balance_tracking_start_date:
        # Get the February 2026 snapshot (the baseline for balance tracking)
        from home.models import MonthEndClose, AccountSnapshot
        tracking_start = account.balance_tracking_start_date
        prev_month = tracking_start.replace(day=1) - timedelta(days=1)
        prev_month_first = prev_month.replace(day=1)

        try:
            month_close = MonthEndClose.objects.get(month=prev_month_first, is_locked=True)
            snapshot = AccountSnapshot.objects.get(month_close=month_close, bank_account=account)
            baseline_balance = snapshot.balance
        except (MonthEndClose.DoesNotExist, AccountSnapshot.DoesNotExist):
            # Fallback if no snapshot found
            baseline_balance = account.current_balance

        # The baseline is the balance BEFORE the tracking start date
        # So we use it as the starting point for all balance calculations
        starting_balance = baseline_balance
    else:
        starting_balance = account.current_balance

    # Calculate running balance for each entry (only for tracked period)
    running_balance = starting_balance
    for entry in entries:
        if account.balance_tracking_enabled and account.balance_tracking_start_date:
            if entry["date"] >= account.balance_tracking_start_date:
                running_balance += entry["signed_amount"]
                entry["balance"] = running_balance
            else:
                entry["balance"] = None  # No balance for pre-tracking transactions
        else:
            running_balance += entry["signed_amount"]
            entry["balance"] = running_balance

    # Group entries by month for display (still in chronological order)
    from itertools import groupby
    entries_by_month = []
    for month_key, month_entries in groupby(entries, key=lambda e: e["date"].strftime("%Y-%m")):
        month_entries_list = list(month_entries)
        month_date = datetime.strptime(month_key, "%Y-%m").date()
        entries_by_month.append({
            "month": month_date.strftime("%B %Y"),
            "month_key": month_key,
            "entries": month_entries_list,
            "month_inflow": sum(e["raw_amount"] for e in month_entries_list if e["is_inflow"]),
            "month_outflow": sum(e["raw_amount"] for e in month_entries_list if not e["is_inflow"]),
            "month_net": sum(e["signed_amount"] for e in month_entries_list),
        })

    # Now reverse everything to show most recent month first (with most recent transactions first within each month)
    entries_by_month.reverse()
    for month_group in entries_by_month:
        month_group["entries"].reverse()

    # --- Overall summaries ---
    total_inflow = sum(e["raw_amount"] for e in entries if e["is_inflow"])
    total_outflow = sum(e["raw_amount"] for e in entries if not e["is_inflow"])
    net_change = sum(e["signed_amount"] for e in entries)

    # Ensure Decimals for template math/formatting
    total_inflow = total_inflow or Decimal("0.00")
    total_outflow = total_outflow or Decimal("0.00")
    net_change = net_change or Decimal("0.00")

    # Ending balance is the balance after the last tracked transaction in the period
    # Find the last entry with a balance (in reverse order since entries are now most recent first)
    ending_balance = account.current_balance
    for entry in entries:
        if entry.get("balance") is not None:
            ending_balance = entry["balance"]
            break  # Found the most recent entry with a balance

    context = {
        "account": account,
        "entries_by_month": entries_by_month,
        "start_date": first_day,
        "end_date": last_day,
        "date_range_display": date_range_display,
        "starting_balance": starting_balance,
        "ending_balance": ending_balance,
        "total_inflow": total_inflow,
        "total_outflow": total_outflow,
        "net_change": net_change,
    }

    return render(request, "bank_account_detail.html", context)


@require_POST
def create_balance_adjustment(request, account_id):
    """
    Create a manual balance adjustment for an account.

    Used for bank reconciliation when there are differences between
    the tracked balance and the actual bank statement (e.g., bank fees,
    interest earned, or other items not recorded as transactions).
    """
    account = get_object_or_404(BankAccount, pk=account_id)

    try:
        # Parse form data
        adjustment_date = request.POST.get('adjustment_date')
        amount = request.POST.get('amount')
        reason = request.POST.get('reason', '')
        notes = request.POST.get('notes', '')

        # Validate required fields
        if not adjustment_date or not amount:
            messages.error(request, "Date and amount are required.")
            return redirect('bank_account_detail', account_id=account_id)

        # Parse date
        try:
            parsed_date = datetime.strptime(adjustment_date, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, "Invalid date format. Use YYYY-MM-DD.")
            return redirect('bank_account_detail', account_id=account_id)

        # Parse amount
        try:
            parsed_amount = Decimal(amount)
        except (InvalidOperation, ValueError):
            messages.error(request, "Invalid amount.")
            return redirect('bank_account_detail', account_id=account_id)

        # Create the adjustment
        adjustment = BalanceAdjustment.objects.create(
            bank_account=account,
            date=parsed_date,
            amount=parsed_amount,
            reason=reason or 'Manual adjustment',
            notes=notes,
            created_by=request.user.username if request.user.is_authenticated else 'System'
        )

        sign = '+' if parsed_amount >= 0 else ''
        messages.success(
            request,
            f"Balance adjustment created: {sign}${parsed_amount} on {parsed_date} ({reason})"
        )

    except Exception as e:
        messages.error(request, f"Error creating balance adjustment: {str(e)}")

    return redirect('bank_account_detail', account_id=account_id)


@require_http_methods(["GET", "POST"])
def unassigned_transactions(request):
    """
    Helper page (not linked in main UI) to clean up:
      - Income/Expense rows missing a bank_account
      - Historical reclassification of some expenses as transfers
      - Withholding bucket assignment on transfers
      - Targeted Foxview Insurance cleanup.

    IMPORTANT: We never change rows that already have a bank_account or bucket
    except where explicitly requested by these helper actions.
    """
    accounts = BankAccount.objects.all().order_by("name")

    def build_unassigned_context(extra=None):
        incomes = (
            Income.objects
            .filter(bank_account__isnull=True)
            .order_by("-date", "-id")
        )
        expenses = (
            Expense.objects
            .filter(bank_account__isnull=True)
            .select_related("category")
            .order_by("-date", "-id")
        )
        ctx = {
            "incomes": incomes,
            "expenses": expenses,
            "accounts": accounts,
        }
        if extra:
            ctx.update(extra)
        return ctx

    if request.method == "POST":

        # --------------------------------------------------
        # 1) Auto-assign bank accounts using rules
        # --------------------------------------------------
        if "auto_assign" in request.POST:
            td_chequings = BankAccount.objects.filter(
                name__iexact="TD CHEQUINGS"
            ).first()
            td_aeroplan_visa = BankAccount.objects.filter(
                name__iexact="TD AEROPLAN VISA"
            ).first()

            if not td_chequings:
                messages.error(
                    request,
                    "Auto-assign: No bank account found with name 'TD CHEQUINGS' "
                    "(case-insensitive). Please create or rename it, then try again.",
                )
            if not td_aeroplan_visa:
                messages.error(
                    request,
                    "Auto-assign: No bank account found with name 'TD AEROPLAN VISA' "
                    "(case-insensitive). Please create or rename it, then try again.",
                )

            assigned_incomes = 0
            assigned_expenses = 0

            # Income rule: all unassigned incomes → TD CHEQUINGS
            if td_chequings:
                for inc in Income.objects.filter(bank_account__isnull=True):
                    inc.bank_account = td_chequings
                    inc.save(update_fields=["bank_account"])
                    assigned_incomes += 1

            # Expense rules
            # These categories are handled by the transfer reclassification step, so skip here.
            skip_cat_keys = {
                "arnprior property tax",
                "rrsp contributions",
                "foxview down payment savings",
                "arnprior rental tax withholding (loft)",
                "arnprior rental tax withholding (main)",
            }

            chequings_cat_keys = {
                "arnprior insurance",
                "foxview property tax",
                "arnprior snow removal",
                "arnprior mortgage interest",
                "arnprior mortgage principal",
                "foxview hydro",
                "foxview insurance",
                "foxview internet",
                "arnprior hydro",
            }

            aeroplan_visa_cat_keys = {
                "gas",
                "business expense",
                "miscellaneous",
            }

            expense_qs = (
                Expense.objects
                .filter(bank_account__isnull=True)
                .select_related("category")
            )

            for exp in expense_qs:
                if not exp.category_id:
                    continue

                cat_name = (exp.category.name or "").strip()
                cat_key = cat_name.lower()
                vendor = (exp.vendor_name or "").strip()
                vendor_lower = vendor.lower()

                # Skip anything that will be covered by the transfer reclassification step
                if cat_key in skip_cat_keys:
                    continue

                target_account = None

                # Groceries rule:
                # - Groceries are on TD AEROPLAN VISA
                # - unless vendor is Costco → TD CHEQUINGS
                if cat_key == "groceries":
                    if "costco" in vendor_lower and td_chequings:
                        target_account = td_chequings
                    elif td_aeroplan_visa:
                        target_account = td_aeroplan_visa

                elif cat_key in chequings_cat_keys and td_chequings:
                    target_account = td_chequings

                elif cat_key in aeroplan_visa_cat_keys and td_aeroplan_visa:
                    target_account = td_aeroplan_visa

                if not target_account:
                    continue

                exp.bank_account = target_account
                exp.save(update_fields=["bank_account"])
                assigned_expenses += 1

            if assigned_incomes or assigned_expenses:
                messages.success(
                    request,
                    f"Auto-assign complete: set bank accounts for "
                    f"{assigned_incomes} income(s) and {assigned_expenses} expense(s). "
                    f"Transactions already linked to an account were not changed.",
                )
            else:
                messages.info(
                    request,
                    "Auto-assign finished but did not change any rows. "
                    "Either there were no unassigned transactions, or account names/rules did not match.",
                )

            return redirect("unassigned_transactions")

        # --------------------------------------------------
        # 2) Reclassify some expenses as transfers + fix others
        # --------------------------------------------------
        if "reclassify_transfers" in request.POST:
            td_chequings = BankAccount.objects.filter(
                name__iexact="TD CHEQUINGS"
            ).first()
            td_visa = BankAccount.objects.filter(
                name__iexact="TD AEROPLAN VISA"
            ).first()
            ws_cash = BankAccount.objects.filter(
                name__iexact="Wealthsimple Cash"
            ).first()
            ws_rrsp = BankAccount.objects.filter(
                name__iexact="Wealthsimple RRSP"
            ).first()

            missing = []
            if not td_chequings:
                missing.append("TD CHEQUINGS")
            if not ws_cash:
                missing.append("Wealthsimple Cash")
            if not ws_rrsp:
                missing.append("Wealthsimple RRSP")
            if not td_visa:
                missing.append("TD AEROPLAN VISA")

            if missing:
                messages.error(
                    request,
                    "Reclassify: Could not find required account(s) "
                    + ", ".join(f"'{m}'" for m in missing)
                    + " (case-insensitive). Please create/rename them and try again.",
                )
                return redirect("unassigned_transactions")

            created_transfers = 0
            assigned_chequings = 0
            assigned_visa = 0

            THRESHOLD = Decimal("550.00")

            expenses = (
                Expense.objects
                .filter(bank_account__isnull=True)
                .select_related("category")
            )

            for exp in expenses:
                if not exp.category:
                    continue

                cat_name = (exp.category.name or "").strip()
                cat_key = cat_name.lower()
                amt = exp.amount

                def make_transfer(to_account):
                    nonlocal created_transfers
                    Transfer.objects.create(
                        date=exp.date,
                        amount=amt,
                        from_account=td_chequings,
                        to_account=to_account,
                        description=f"Reclassified from expense: {cat_name}",
                        notes=exp.notes or exp.vendor_name or "",
                        # withholding_category will be added later by the helper actions
                    )
                    exp.delete()
                    created_transfers += 1

                # ----------------
                # TRANSFER RULES
                # ----------------

                # 1) Arnprior Property Tax – thresholded
                if cat_key == "arnprior property tax":
                    if amt <= THRESHOLD:
                        make_transfer(ws_cash)
                    # > 550 → leave unassigned for manual review
                    continue

                # 2) RRSP Contributions – always Chequings → WS RRSP
                if cat_key == "rrsp contributions":
                    make_transfer(ws_rrsp)
                    continue

                # 3) Foxview Insurance exactly 250 → transfer Chequings → WS Cash
                if cat_key == "foxview insurance":
                    if amt == Decimal("250.00"):
                        make_transfer(ws_cash)
                    # other amounts stay for manual handling / other tools
                    continue

                # 4/5/6) Foxview Down Payment Savings, Arnprior Rental Tax Withholding (LOFT/MAIN)
                #        – always Chequings → WS Cash (no threshold)
                if cat_key in {
                    "foxview down payment savings",
                    "arnprior rental tax withholding (loft)",
                    "arnprior rental tax withholding (main)",
                }:
                    make_transfer(ws_cash)
                    continue

                # ----------------
                # REMAIN AS EXPENSE (assign account)
                # ----------------

                # Chequings expenses
                if cat_key in {
                    "arnprior snow removal",
                    "foxview property tax",
                    "subaru insurance",
                    "arnprior heat",
                    "foxview internet",
                }:
                    exp.bank_account = td_chequings
                    exp.save(update_fields=["bank_account"])
                    assigned_chequings += 1
                    continue

                # Visa expenses
                if cat_key in {
                    "cell phone",
                    "arnprior internet",
                    "foxview heat",
                    "digital subscriptions",
                    "restaurants",
                }:
                    exp.bank_account = td_visa
                    exp.save(update_fields=["bank_account"])
                    assigned_visa += 1
                    continue

            messages.success(
                request,
                "Reclassification complete: "
                f"{created_transfers} transfer(s) created, "
                f"{assigned_chequings} expense(s) assigned to TD CHEQUINGS, "
                f"{assigned_visa} expense(s) assigned to TD AEROPLAN VISA."
            )
            return redirect("unassigned_transactions")

        # --------------------------------------------------
        # 3a) PREVIEW: Attach withholding buckets to reclassified transfers
        # --------------------------------------------------
        if "assign_transfer_buckets_preview" in request.POST:
            # Buckets: use your exact WithholdingCategory names
            bucket_arnprior_prop = WithholdingCategory.objects.filter(
                name__iexact="Arnprior Property Tax"
            ).first()
            bucket_arnprior_rental = WithholdingCategory.objects.filter(
                name__iexact="Arnprior Rental Income Tax"
            ).first()
            bucket_foxview_ins = WithholdingCategory.objects.filter(
                name__iexact="Foxview Insurance"
            ).first()

            missing_buckets = []
            if not bucket_arnprior_prop:
                missing_buckets.append("Arnprior Property Tax")
            if not bucket_arnprior_rental:
                missing_buckets.append("Arnprior Rental Income Tax")
            if not bucket_foxview_ins:
                missing_buckets.append("Foxview Insurance")

            if missing_buckets:
                messages.error(
                    request,
                    "Assign buckets: Could not find required withholding bucket(s) "
                    + ", ".join(f"'{b}'" for b in missing_buckets)
                    + ". Please create/rename them and try again.",
                )
                return redirect("unassigned_transactions")

            transfers = Transfer.objects.filter(
                withholding_category__isnull=True,
                description__startswith="Reclassified from expense:",
            ).order_by("date", "id")

            preview_rows = []
            for t in transfers:
                desc = t.description or ""
                proposed_bucket = None
                rule = ""

                if "Arnprior Property Tax" in desc:
                    proposed_bucket = bucket_arnprior_prop
                    rule = "Arnprior Property Tax → Arnprior Property Tax bucket"
                elif "Arnprior Rental Tax Withholding (MAIN)" in desc or "Arnprior Rental Tax Withholding (LOFT)" in desc:
                    proposed_bucket = bucket_arnprior_rental
                    rule = "Rental Tax Withholding (MAIN/LOFT) → Arnprior Rental Income Tax bucket"
                elif "Foxview Insurance" in desc:
                    proposed_bucket = bucket_foxview_ins
                    rule = "Foxview Insurance → Foxview Insurance bucket"
                else:
                    rule = "No matching rule (will be skipped)"

                preview_rows.append(
                    {
                        "transfer": t,
                        "proposed_bucket": proposed_bucket,
                        "rule": rule,
                        "will_change": proposed_bucket is not None,
                    }
                )

            context = build_unassigned_context(
                {"bucket_preview": preview_rows}
            )
            return render(request, "unassigned_transactions.html", context)

        # --------------------------------------------------
        # 3b) APPLY: Attach withholding buckets to reclassified transfers
        # --------------------------------------------------
        if "assign_transfer_buckets_apply" in request.POST:
            bucket_arnprior_prop = WithholdingCategory.objects.filter(
                name__iexact="Arnprior Property Tax"
            ).first()
            bucket_arnprior_rental = WithholdingCategory.objects.filter(
                name__iexact="Arnprior Rental Income Tax"
            ).first()
            bucket_foxview_ins = WithholdingCategory.objects.filter(
                name__iexact="Foxview Insurance"
            ).first()

            missing_buckets = []
            if not bucket_arnprior_prop:
                missing_buckets.append("Arnprior Property Tax")
            if not bucket_arnprior_rental:
                missing_buckets.append("Arnprior Rental Income Tax")
            if not bucket_foxview_ins:
                missing_buckets.append("Foxview Insurance")

            if missing_buckets:
                messages.error(
                    request,
                    "Assign buckets: Could not find required withholding bucket(s) "
                    + ", ".join(f"'{b}'" for b in missing_buckets)
                    + ". Please create/rename them and try again.",
                )
                return redirect("unassigned_transactions")

            updated = 0
            skipped = 0

            transfers = Transfer.objects.filter(
                withholding_category__isnull=True,
                description__startswith="Reclassified from expense:",
            )

            for t in transfers:
                desc = t.description or ""
                if "Arnprior Property Tax" in desc and bucket_arnprior_prop:
                    t.withholding_category = bucket_arnprior_prop
                elif ("Arnprior Rental Tax Withholding (MAIN)" in desc or
                      "Arnprior Rental Tax Withholding (LOFT)" in desc):
                    if bucket_arnprior_rental:
                        t.withholding_category = bucket_arnprior_rental
                    else:
                        skipped += 1
                        continue
                elif "Foxview Insurance" in desc and bucket_foxview_ins:
                    t.withholding_category = bucket_foxview_ins
                else:
                    skipped += 1
                    continue

                t.save(update_fields=["withholding_category"])
                updated += 1

            messages.success(
                request,
                f"Assigned withholding buckets to {updated} transfer(s). "
                f"Skipped {skipped} transfer(s) that did not match a known rule.",
            )
            return redirect("unassigned_transactions")

        # --------------------------------------------------
        # 4) Foxview Insurance cleanup:
        #    - convert 250/500 expenses into transfers
        #    - create missing real expenses from legacy payouts
        # --------------------------------------------------
        if "fix_foxview_insurance" in request.POST:
            td_chequings = BankAccount.objects.filter(
                name__iexact="TD CHEQUINGS"
            ).first()
            ws_cash = BankAccount.objects.filter(
                name__iexact="Wealthsimple Cash"
            ).first()
            foxview_cat = Category.objects.filter(
                name__iexact="Foxview Insurance"
            ).first()
            foxview_bucket = WithholdingCategory.objects.filter(
                name__iexact="Foxview Insurance"
            ).first()

            missing = []
            if not td_chequings:
                missing.append("TD CHEQUINGS")
            if not ws_cash:
                missing.append("Wealthsimple Cash")
            if not foxview_cat:
                missing.append("Foxview Insurance (expense category)")
            if not foxview_bucket:
                missing.append("Foxview Insurance (withholding bucket)")

            if missing:
                messages.error(
                    request,
                    "Foxview Insurance cleanup: Missing required objects: "
                    + ", ".join(f"'{m}'" for m in missing)
                    + ". Please create/rename them and try again.",
                )
                return redirect("unassigned_transactions")

            # 4a) Convert contribution expenses (250 / 500) into transfers
            from django.db.models import Q

            contrib_qs = Expense.objects.filter(
                category=foxview_cat,
                bank_account=td_chequings,
            ).filter(
                Q(amount=Decimal("250.00")) | Q(amount=Decimal("500.00"))
            )

            converted = 0
            for exp in contrib_qs:
                Transfer.objects.create(
                    date=exp.date,
                    amount=exp.amount,
                    from_account=td_chequings,
                    to_account=ws_cash,
                    description=exp.vendor_name or "Foxview Insurance transfer",
                    notes=(exp.notes or "").strip() or f"Reclassified from expense id {exp.id}",
                    withholding_category=foxview_bucket,
                )
                exp.delete()
                converted += 1

            # 4b) Create missing real expenses from legacy payouts (negative ledger entries)
            created_expenses = 0
            if foxview_bucket:
                payout_txs = foxview_bucket.transactions.filter(amount__lt=0)

                for tx in payout_txs:
                    amt = -tx.amount  # make positive
                    exists = Expense.objects.filter(
                        category=foxview_cat,
                        amount=amt,
                        date=tx.date,
                        withholding_category=foxview_bucket,
                    ).exists()
                    if exists:
                        continue

                    Expense.objects.create(
                        date=tx.date,
                        vendor_name=tx.note or "Foxview Insurance bill",
                        category=foxview_cat,
                        amount=amt,
                        bank_account=td_chequings,
                        location="Ottawa",
                        withholding_category=foxview_bucket,
                        notes="Created from Foxview Insurance withholding ledger payout",
                    )
                    created_expenses += 1

            messages.success(
                request,
                "Foxview Insurance cleanup complete: "
                f"{converted} contribution expense(s) converted to transfers, "
                f"{created_expenses} bill expense(s) created from legacy ledger."
            )
            return redirect("unassigned_transactions")

        # --------------------------------------------------
        # 5) Manual single-income assignment
        # --------------------------------------------------
        if "assign_income_id" in request.POST:
            income = get_object_or_404(Income, pk=request.POST["assign_income_id"])
            bank_account_id = (request.POST.get("bank_account") or "").strip()
            if not bank_account_id:
                messages.error(request, "Please choose a bank account before saving this income.")
            else:
                income.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                income.save(update_fields=["bank_account"])
                messages.success(
                    request,
                    f"Updated income on {income.date} for ${income.amount} "
                    f"to use account '{income.bank_account.name}'."
                )
            return redirect("unassigned_transactions")

        # --------------------------------------------------
        # 6) Manual single-expense assignment
        # --------------------------------------------------
        if "assign_expense_id" in request.POST:
            expense = get_object_or_404(Expense, pk=request.POST["assign_expense_id"])
            bank_account_id = (request.POST.get("bank_account") or "").strip()
            if not bank_account_id:
                messages.error(request, "Please choose a bank account before saving this expense.")
            else:
                expense.bank_account = get_object_or_404(BankAccount, pk=bank_account_id)
                expense.save(update_fields=["bank_account"])
                messages.success(
                    request,
                    f"Updated expense on {expense.date} ({expense.vendor_name}) "
                    f"to use account '{expense.bank_account.name}'."
                )
            return redirect("unassigned_transactions")

    # GET (or fallthrough)
    context = build_unassigned_context()
    return render(request, "unassigned_transactions.html", context)

def import_batch_detail(request, batch_id):
    batch = get_object_or_404(ImportBatch, pk=batch_id)
    expenses = Expense.objects.filter(import_batch=batch).order_by("-date")
    incomes = Income.objects.filter(import_batch=batch).order_by("-date")

    return render(request, "import_batch_detail.html", {
        "batch": batch,
        "expenses": expenses,
        "incomes": incomes,
    })

@require_http_methods(["GET", "POST"])
def import_transactions(request):
    if request.method == "GET":
        upload_form = CSVUploadForm()
        recent_batches = ImportBatch.objects.select_related("bank_account").all()[:10]


        return render(request, "import_transactions.html", {
            "step": "upload",
            "upload_form": upload_form,
            "recent_batches": recent_batches,
        })

    step = request.POST.get("step", "upload")

    if step == "upload" and request.FILES.get("csv_file"):
        upload_form = CSVUploadForm(request.POST, request.FILES)
        recent_batches = ImportBatch.objects.select_related("bank_account").all()[:10]

        if not upload_form.is_valid():
            return render(request, "import_transactions.html", {
                "step": "upload",
                "upload_form": upload_form,
                "recent_batches": recent_batches,
            })

        csv_file = upload_form.cleaned_data["csv_file"]
        bank_account = upload_form.cleaned_data["bank_account"]
        uploaded_filename = csv_file.name

        try:
            decoded = io.TextIOWrapper(csv_file.file, encoding="utf-8")
        except Exception:
            decoded = io.TextIOWrapper(csv_file.file, encoding="latin-1")

        reader = csv.reader(decoded)

        initial_rows = []
        category_cache = {}
        missing_categories = set()
        hydro_candidates = []
        earliest_date_in_file = None
        latest_date_in_file = None

        income_cat_cache = {c.name: c for c in IncomeCategory.objects.all()}
        arnprior_shared_unit_id = get_arnprior_shared_unit_id()
        foxview_shared_unit_id = get_foxview_shared_unit_id()

        for row in reader:
            if not row or all(not cell.strip() for cell in row):
                continue

            raw_date = row[0].strip() if len(row) > 0 else ""
            raw_desc = row[1].strip() if len(row) > 1 else ""
            raw_withdrawal = row[2].strip() if len(row) > 2 else ""
            raw_deposit = row[3].strip() if len(row) > 3 else ""

            desc_upper = (raw_desc or "").upper()

            if "TFR-TO C/C" in desc_upper:
                continue

            if not raw_desc and not raw_withdrawal and not raw_deposit:
                continue

            parsed_date = None
            for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                try:
                    parsed_date = datetime.strptime(raw_date, fmt).date()
                    break
                except Exception:
                    continue

            if parsed_date is None:
                continue

            if earliest_date_in_file is None or parsed_date < earliest_date_in_file:
                earliest_date_in_file = parsed_date
            if latest_date_in_file is None or parsed_date > latest_date_in_file:
                latest_date_in_file = parsed_date

            amount_str = None
            entry_type_default = "expense"

            if raw_withdrawal and not raw_deposit:
                amount_str = raw_withdrawal
                entry_type_default = "expense"
            elif raw_deposit and not raw_withdrawal:
                amount_str = raw_deposit
                entry_type_default = "income"
            elif raw_withdrawal and raw_deposit:
                amount_str = raw_withdrawal
                entry_type_default = "expense"
            else:
                continue

            amount_str = amount_str.replace(",", "")
            try:
                amount = Decimal(amount_str)
            except Exception:
                continue

            amount = abs(amount)

            entry_type, income_source_name = apply_income_rules(desc_upper, amount, entry_type_default, parsed_date=parsed_date)


            income_source_obj = None
            if entry_type == "income" and income_source_name:
                income_source_obj = income_cat_cache.get(income_source_name)
                if income_source_obj is None:
                    income_source_obj, _ = IncomeCategory.objects.get_or_create(name=income_source_name)
                    income_cat_cache[income_source_name] = income_source_obj

            expense_category = None
            if entry_type == "expense":
                expense_category = apply_expense_rules(desc_upper, amount, category_cache, missing_categories)

            if "HYDRO ONE" in desc_upper and entry_type == "expense":
                hydro_candidates.append((len(initial_rows), amount))

            expense_rental_unit_id = None
            if entry_type == "expense" and expense_category:
                category_upper = (expense_category.name or "").upper()

                if "ARNPRIOR" in category_upper:
                    if arnprior_shared_unit_id:
                        expense_rental_unit_id = arnprior_shared_unit_id

                elif "FOXVIEW" in category_upper:
                    if foxview_shared_unit_id:
                        expense_rental_unit_id = foxview_shared_unit_id

            initial_rows.append({
                "entry_type": entry_type,
                "date": parsed_date,
                "vendor_name": raw_desc or "Unknown Vendor",
                "amount": amount,
                "location": "Ottawa",
                "notes": "",
                "expense_category": expense_category,
                "income_source": income_source_obj,
                "income_rental_unit": (
                    income_source_obj.default_rental_unit_id if income_source_obj and income_source_obj.default_rental_unit_id else None),
                "apply_to_withholding": False,
                "is_withholding_payout": False,
                "withholding_category": None,
                "expense_rental_unit": expense_rental_unit_id,
            })

        if bank_account and earliest_date_in_file and latest_date_in_file:
            overlapping = ImportBatch.objects.filter(
                bank_account=bank_account,
                earliest_date__lte=latest_date_in_file,
                latest_date__gte=earliest_date_in_file,
            )
            if overlapping.exists():
                ranges = "; ".join(f"{b.earliest_date} to {b.latest_date}" for b in overlapping)
                messages.warning(
                    request,
                    f"This CSV covers {earliest_date_in_file} to {latest_date_in_file}, "
                    f"which overlaps with existing imports for this account: {ranges}. "
                    f"Duplicates will be skipped where detected."
                )

        if hydro_candidates:
            if len(hydro_candidates) == 1:
                idx, amt = hydro_candidates[0]
                cat_name = "Foxview Hydro" if amt > Decimal("200.00") else "Arnprior Hydro"
                cat = get_category_cached(cat_name, category_cache, missing_categories)
                if cat:
                    initial_rows[idx]["expense_category"] = cat
            else:
                sorted_by_amount = sorted(hydro_candidates, key=lambda x: x[1])
                for idx, amt in sorted_by_amount:
                    if len(sorted_by_amount) == 2:
                        cat_name = "Arnprior Hydro" if (idx, amt) == sorted_by_amount[0] else "Foxview Hydro"
                    else:
                        cat_name = "Foxview Hydro" if amt > Decimal("200.00") else "Arnprior Hydro"
                    cat = get_category_cached(cat_name, category_cache, missing_categories)
                    if cat:
                        initial_rows[idx]["expense_category"] = cat

        if not initial_rows:
            messages.warning(request, "No valid transactions were found in the CSV (check the file format).")
            upload_form = CSVUploadForm()
            return render(request, "import_transactions.html", {
                "step": "upload",
                "upload_form": upload_form,
                "recent_batches": ImportBatch.objects.select_related("bank_account").all()[:10],
            })

        if missing_categories:
            missing_list = ", ".join(sorted(missing_categories))
            messages.warning(
                request,
                f"The following auto-mapped categories were not found in your database and were skipped: {missing_list}."
            )

        formset = TransactionImportFormSet(initial=initial_rows)
        income_rental_unit_map = build_income_rental_unit_map()

        return render(request, "import_transactions.html", {
            "step": "review",
            "formset": formset,
            "selected_bank_account": bank_account,
            "uploaded_filename": uploaded_filename,
            "income_rental_unit_map": income_rental_unit_map,
        })

    if step == "review":
        formset = TransactionImportFormSet(request.POST)

        bank_account = None
        bank_account_id = request.POST.get("bank_account_id")
        if bank_account_id:
            try:
                bank_account = BankAccount.objects.get(pk=bank_account_id)
            except BankAccount.DoesNotExist:
                bank_account = None

        uploaded_filename = request.POST.get("uploaded_filename", "")

        if not formset.is_valid():
            messages.error(request, "There were errors in the form. Please correct them.")
            income_rental_unit_map = build_income_rental_unit_map()


            return render(request, "import_transactions.html", {
                "step": "review",
                "formset": formset,
                "selected_bank_account": bank_account,
                "uploaded_filename": uploaded_filename,
                "income_rental_unit_map": income_rental_unit_map,  # ✅ add this
            })

        created_expenses = 0
        created_incomes = 0
        created_transfers = 0
        skipped_duplicates = 0
        created_withholding_transactions = 0

        earliest_date = None
        latest_date = None
        total_expense_amount = Decimal("0.00")
        total_income_amount = Decimal("0.00")

        expense_objs = []
        income_objs = []
        transfer_objs = []
        withholding_txns = []

        seen_expense_keys = set()
        seen_income_keys = set()
        seen_transfer_keys = set()

        for form in formset:
            cd = form.cleaned_data
            if not cd:
                continue
            if cd.get("skip"):
                continue

            entry_type = cd.get("entry_type")
            date_val = cd.get("date")
            vendor_name = cd.get("vendor_name")
            amount = cd.get("amount")
            location = cd.get("location") or "Ottawa"
            notes = cd.get("notes")
            expense_category = cd.get("expense_category")
            expense_rental_unit = cd.get("expense_rental_unit")
            income_source = cd.get("income_source")
            income_rental_unit = cd.get("income_rental_unit")
            apply_to_withholding = cd.get("apply_to_withholding")
            is_withholding_payout = cd.get("is_withholding_payout")
            withholding_category = cd.get("withholding_category")

            if not (date_val and amount):
                continue

            if earliest_date is None or date_val < earliest_date:
                earliest_date = date_val
            if latest_date is None or date_val > latest_date:
                latest_date = date_val

            if entry_type == "expense":
                if is_withholding_payout and withholding_category:
                    payout_amount = -amount if amount > 0 else amount
                    wt = WithholdingTransaction(
                        category=withholding_category,
                        date=date_val,
                        amount=payout_amount,
                        note=notes or vendor_name or "",
                    )
                    withholding_txns.append(wt)
                    created_withholding_transactions += 1
                    continue

                if not (vendor_name and expense_category):
                    continue

                exp_key = (date_val, vendor_name, amount, expense_category.id)

                if exp_key in seen_expense_keys or Expense.objects.filter(
                    date=date_val,
                    vendor_name=vendor_name,
                    amount=amount,
                    category=expense_category,
                ).exists():
                    skipped_duplicates += 1
                    continue

                seen_expense_keys.add(exp_key)

                exp = Expense(
                    date=date_val,
                    vendor_name=vendor_name,
                    amount=amount,
                    category=expense_category,
                    location=location,
                    notes=notes or "",
                    bank_account=bank_account,
                    rental_unit=expense_rental_unit,
                )
                expense_objs.append(exp)
                total_expense_amount += amount
                created_expenses += 1

                if apply_to_withholding and withholding_category:
                    wt = WithholdingTransaction(
                        category=withholding_category,
                        date=date_val,
                        amount=amount,
                        note=notes or vendor_name or "",
                    )
                    withholding_txns.append(wt)
                    created_withholding_transactions += 1

            elif entry_type == "income":
                if not income_source:
                    continue

                inc_key = (date_val, amount, income_source.id)

                if inc_key in seen_income_keys or Income.objects.filter(
                    date=date_val,
                    amount=amount,
                    income_category=income_source,
                ).exists():
                    skipped_duplicates += 1
                    continue

                seen_income_keys.add(inc_key)

                inc = Income(
                    date=date_val,
                    amount=amount,
                    income_category=income_source,
                    category=income_source.name,
                    taxable=income_source.taxable_default,
                    notes=notes or "",
                    bank_account=bank_account,
                )

                # If user selected an override, use it; otherwise fall back to the IncomeCategory default
                if income_rental_unit:
                    inc.rental_unit = income_rental_unit
                elif income_source and income_source.default_rental_unit_id:
                    inc.rental_unit = income_source.default_rental_unit

                income_objs.append(inc)
                total_income_amount += amount
                created_incomes += 1

            elif entry_type == "transfer":
                from_account = cd.get("from_account")
                to_account = cd.get("to_account")
                withholding_category = cd.get("withholding_category")

                # Validation: at least one account (should already be validated by form)
                if not from_account and not to_account:
                    continue

                # Create duplicate detection key
                transfer_key = (
                    date_val,
                    from_account.id if from_account else None,
                    to_account.id if to_account else None,
                    amount
                )

                # Check for duplicates in current batch
                if transfer_key in seen_transfer_keys:
                    skipped_duplicates += 1
                    continue

                # Check database for existing transfer
                transfer_exists = Transfer.objects.filter(
                    date=date_val,
                    amount=amount,
                )
                if from_account:
                    transfer_exists = transfer_exists.filter(from_account=from_account)
                if to_account:
                    transfer_exists = transfer_exists.filter(to_account=to_account)

                if transfer_exists.exists():
                    skipped_duplicates += 1
                    continue

                seen_transfer_keys.add(transfer_key)

                # Create Transfer object
                transfer = Transfer(
                    date=date_val,
                    amount=amount,
                    description=vendor_name or "",
                    notes=notes or "",
                    from_account=from_account,
                    to_account=to_account,
                    withholding_category=withholding_category,
                )
                transfer_objs.append(transfer)
                created_transfers += 1

        total_transactions = created_expenses + created_incomes + created_transfers

        if total_transactions > 0 and earliest_date and latest_date:
            batch = ImportBatch.objects.create(
                bank_account=bank_account,
                earliest_date=earliest_date,
                latest_date=latest_date,
                total_transactions=total_transactions,
                total_income_amount=total_income_amount,
                total_expense_amount=total_expense_amount,
                filename=uploaded_filename or "",
            )

            for exp in expense_objs:
                exp.import_batch = batch
                exp.save()
            for inc in income_objs:
                inc.import_batch = batch
                inc.save()
            for transfer in transfer_objs:
                transfer.import_batch = batch
                transfer.save()
        else:
            for exp in expense_objs:
                exp.save()
            for inc in income_objs:
                inc.save()
            for transfer in transfer_objs:
                transfer.save()

        for wt in withholding_txns:
            wt.save()

        msg = f"Imported {created_expenses} expense(s), {created_incomes} income, and {created_transfers} transfer transaction(s)."
        if created_withholding_transactions:
            msg += f" Applied {created_withholding_transactions} withholding bucket adjustment(s)."
        if skipped_duplicates:
            msg += f" Skipped {skipped_duplicates} duplicate(s)."
        messages.success(request, msg)

        return redirect("dashboard")

    upload_form = CSVUploadForm()
    recent_batches = ImportBatch.objects.select_related("bank_account").all()[:10]
    return render(request, "import_transactions.html", {
        "step": "upload",
        "upload_form": upload_form,
        "recent_batches": recent_batches,
    })

@csrf_exempt
def update_expense(request):
    """
    Legacy endpoint still referenced in urls.py.
    Keep it working (and updated) even if the dashboard modal is the main flow.
    """
    if request.method != "POST":
        return HttpResponseBadRequest("Invalid method")

    expense_id = request.POST.get("id")
    try:
        expense = Expense.objects.get(id=expense_id)
    except Expense.DoesNotExist:
        return HttpResponseBadRequest("Expense not found")

    if "delete" in request.POST:
        expense.delete()
    else:
        expense.date = request.POST.get("date")
        expense.vendor_name = request.POST.get("vendor_name")
        expense.category = get_object_or_404(Category, id=request.POST.get("category_id"))
        expense.location = request.POST.get("location", "Ottawa")
        expense.amount = request.POST.get("amount")
        expense.notes = request.POST.get("notes", "")

        # ✅ Optional rental fields if caller sends them
        rental_unit_id = (request.POST.get("rental_unit") or "").strip()
        expense.rental_unit = get_object_or_404(RentalUnit, pk=rental_unit_id) if rental_unit_id else None

        cra_category_id = (request.POST.get("cra_category") or "").strip()
        expense.cra_category = get_object_or_404(CRARentalExpenseCategory, pk=cra_category_id) if cra_category_id else None

        pct_str = (request.POST.get("rental_business_use_pct") or "").strip()
        expense.rental_business_use_pct = Decimal(pct_str) if pct_str else None

        expense.save()

    selected_month = request.GET.get("month") or expense.date.strftime("%Y-%m")
    return redirect(f"/?month={selected_month}")

def withholding_overview(request):
    """
    Overview of withholding accounts and their buckets.

    Balances and monthly activity are fully derived from:
    - Transfers linked to a withholding bucket
    - Expenses funded from a withholding bucket
    """
    # Handle POST - Create new bucket
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_bucket":
            account_id = request.POST.get("account")
            name = request.POST.get("name", "").strip()
            monthly_target = request.POST.get("monthly_target", "").strip()
            target_amount = request.POST.get("target_amount", "").strip()
            next_due_date = request.POST.get("next_due_date", "").strip()

            if account_id and name:
                try:
                    account = BankAccount.objects.get(id=account_id, is_withholding_account=True)

                    # Create the bucket
                    bucket = WithholdingCategory(
                        account=account,
                        name=name,
                    )

                    # Set optional fields
                    if monthly_target:
                        bucket.monthly_target = Decimal(monthly_target)
                    if target_amount:
                        bucket.target_amount = Decimal(target_amount)
                    if next_due_date:
                        bucket.next_due_date = next_due_date

                    bucket.save()
                    messages.success(request, f"Created new bucket: {name}")
                    return redirect("withholding_overview")

                except BankAccount.DoesNotExist:
                    messages.error(request, "Invalid withholding account selected.")
                except Exception as e:
                    messages.error(request, f"Error creating bucket: {str(e)}")
            else:
                messages.error(request, "Account and bucket name are required.")

    # Determine selected month
    today = date.today()
    month_param = (request.GET.get("month") or "").strip()
    if month_param:
        try:
            year, month = map(int, month_param.split("-"))
            month_start = date(year, month, 1)
        except ValueError:
            month_start = date(today.year, today.month, 1)
    else:
        month_start = date(today.year, today.month, 1)

    _, last_day = monthrange(month_start.year, month_start.month)
    month_end = date(month_start.year, month_start.month, last_day)

    selected_month = month_start.strftime("%Y-%m")
    selected_month_display = month_start.strftime("%B %Y")

    # Get withholding accounts and their buckets
    accounts = (
        BankAccount.objects
        .filter(is_withholding_account=True, is_active=True)
        .prefetch_related("withholding_categories")
        .order_by("name")
    )

    # Collect bucket IDs
    bucket_ids = []
    for account in accounts:
        for bucket in account.withholding_categories.all():
            if bucket.id is not None:
                bucket_ids.append(bucket.id)

    # Get withholding accounts for the modal dropdown
    withholding_accounts = BankAccount.objects.filter(
        is_withholding_account=True,
        is_active=True
    ).order_by("name")

    if not bucket_ids:
        return render(
            request,
            "withholding_overview.html",
            {
                "accounts": accounts,
                "selected_month": selected_month,
                "selected_month_display": selected_month_display,
                "withholding_accounts": withholding_accounts,
            },
        )

    # --- Aggregate all-time and monthly activity per bucket ---

    # Transfers (all-time)
    transfers_all = (
        Transfer.objects
        .filter(withholding_category_id__in=bucket_ids)
        .values("withholding_category_id")
        .annotate(
            in_total=Sum(
                Case(
                    When(
                        to_account_id=F("withholding_category__account_id"),
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            out_total=Sum(
                Case(
                    When(
                        from_account_id=F("withholding_category__account_id"),
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
        )
    )
    transfers_all_map = {
        row["withholding_category_id"]: row for row in transfers_all
    }

    # Transfers (this month)
    transfers_month = (
        Transfer.objects
        .filter(
            withholding_category_id__in=bucket_ids,
            date__gte=month_start,
            date__lte=month_end,
        )
        .values("withholding_category_id")
        .annotate(
            in_total=Sum(
                Case(
                    When(
                        to_account_id=F("withholding_category__account_id"),
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
            out_total=Sum(
                Case(
                    When(
                        from_account_id=F("withholding_category__account_id"),
                        then=F("amount"),
                    ),
                    default=Value(0),
                    output_field=DecimalField(max_digits=12, decimal_places=2),
                )
            ),
        )
    )
    transfers_month_map = {
        row["withholding_category_id"]: row for row in transfers_month
    }

    # Expenses (all-time)
    expenses_all = (
        Expense.objects
        .filter(withholding_category_id__in=bucket_ids)
        .values("withholding_category_id")
        .annotate(
            exp_total=Sum(
                "amount",
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
    )
    expenses_all_map = {
        row["withholding_category_id"]: row for row in expenses_all
    }

    # Expenses (this month)
    expenses_month = (
        Expense.objects
        .filter(
            withholding_category_id__in=bucket_ids,
            date__gte=month_start,
            date__lte=month_end,
        )
        .values("withholding_category_id")
        .annotate(
            exp_total=Sum(
                "amount",
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )
    )
    expenses_month_map = {
        row["withholding_category_id"]: row for row in expenses_month
    }

    # Build per-bucket summaries and attach them to bucket objects
    for account in accounts:
        account_total_balance = Decimal("0.00")
        account_month_contrib = Decimal("0.00")
        account_month_payout = Decimal("0.00")

        for bucket in account.withholding_categories.all():
            all_tr = transfers_all_map.get(bucket.id, {})
            month_tr = transfers_month_map.get(bucket.id, {})
            all_exp = expenses_all_map.get(bucket.id, {})
            month_exp = expenses_month_map.get(bucket.id, {})

            in_total = all_tr.get("in_total") or Decimal("0.00")
            out_total = all_tr.get("out_total") or Decimal("0.00")
            exp_total = all_exp.get("exp_total") or Decimal("0.00")

            # All-time derived balance for this bucket
            balance = in_total - out_total - exp_total

            month_in = month_tr.get("in_total") or Decimal("0.00")
            month_out = month_tr.get("out_total") or Decimal("0.00")
            month_exp_total = month_exp.get("exp_total") or Decimal("0.00")

            month_contrib = month_in
            month_payout = month_out + month_exp_total
            month_net = month_contrib - month_payout

            # Remaining to target (if target defined)
            remaining_to_target = None
            if getattr(bucket, "target_amount", None):
                remaining_to_target = bucket.target_amount - balance

            # Attach to bucket instance for template
            bucket.derived_balance = balance
            bucket.month_contrib = month_contrib
            bucket.month_payout = month_payout
            bucket.month_net = month_net
            bucket.remaining_to_target = remaining_to_target

            # Accumulate per-account
            account_total_balance += balance
            account_month_contrib += month_contrib
            account_month_payout += month_payout

        account.total_bucket_balance = account_total_balance
        account.total_month_contrib = account_month_contrib
        account.total_month_payout = account_month_payout

    return render(
        request,
        "withholding_overview.html",
        {
            "accounts": accounts,
            "selected_month": selected_month,
            "selected_month_display": selected_month_display,
            "withholding_accounts": withholding_accounts,
        },
    )



def withholding_category_detail(request, pk):
    """
    Detail view for a single withholding bucket with date range filtering.
    Supports inline editing of expenses and transfers.
    """
    category = get_object_or_404(
        WithholdingCategory.objects.select_related("account"),
        pk=pk,
    )

    # Handle POST requests for editing/deleting expenses and transfers
    if request.method == "POST":
        # Handle expense edit/delete
        if "expense_id" in request.POST:
            expense_id = request.POST.get("expense_id")
            expense = get_object_or_404(Expense, pk=expense_id)

            if request.POST.get("delete_expense") == "1":
                expense.delete()
                messages.success(request, "Expense deleted successfully.")
                return redirect("withholding_category_detail", pk=pk)

            # Update expense fields
            date_str = request.POST.get("date")
            if date_str:
                try:
                    expense.date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            expense.vendor_name = request.POST.get("vendor_name", "")
            expense.location = request.POST.get("location", "")
            expense.notes = request.POST.get("notes", "")

            amount_str = request.POST.get("amount")
            if amount_str:
                try:
                    expense.amount = Decimal(amount_str)
                except (ValueError, InvalidOperation):
                    pass

            # Category
            category_name = request.POST.get("category")
            if category_name:
                cat, _ = Category.objects.get_or_create(name=category_name)
                expense.category = cat

            # Bank account
            bank_account_id = request.POST.get("bank_account")
            if bank_account_id:
                expense.bank_account_id = bank_account_id
            else:
                expense.bank_account = None

            # Rental unit
            rental_unit_id = request.POST.get("rental_unit")
            if rental_unit_id:
                expense.rental_unit_id = rental_unit_id
            else:
                expense.rental_unit = None

            # CRA category
            cra_category_id = request.POST.get("cra_category")
            if cra_category_id:
                expense.cra_category_id = cra_category_id
            else:
                expense.cra_category = None

            expense.save()
            messages.success(request, "Expense updated successfully.")
            return redirect("withholding_category_detail", pk=pk)

        # Handle transfer edit/delete
        elif "transfer_id" in request.POST:
            transfer_id = request.POST.get("transfer_id")
            transfer = get_object_or_404(Transfer, pk=transfer_id)

            if request.POST.get("delete_transfer") == "1":
                transfer.delete()
                messages.success(request, "Transfer deleted successfully.")
                return redirect("withholding_category_detail", pk=pk)

            # Update transfer fields
            date_str = request.POST.get("date")
            if date_str:
                try:
                    transfer.date = datetime.strptime(date_str, "%Y-%m-%d").date()
                except ValueError:
                    pass

            amount_str = request.POST.get("amount")
            if amount_str:
                try:
                    transfer.amount = Decimal(amount_str)
                except (ValueError, InvalidOperation):
                    pass

            # From/To accounts
            from_account_id = request.POST.get("from_account")
            if from_account_id:
                transfer.from_account_id = from_account_id
            else:
                transfer.from_account = None

            to_account_id = request.POST.get("to_account")
            if to_account_id:
                transfer.to_account_id = to_account_id
            else:
                transfer.to_account = None

            # Withholding category
            withholding_category_id = request.POST.get("withholding_category")
            if withholding_category_id:
                transfer.withholding_category_id = withholding_category_id
            else:
                transfer.withholding_category = None

            transfer.notes = request.POST.get("notes", "")

            transfer.save()
            messages.success(request, "Transfer updated successfully.")
            return redirect("withholding_category_detail", pk=pk)

    # ---------- Date Range Selection ----------
    today = date.today()
    selected_range = request.GET.get("range", "12")  # Default: Last 12 months

    # Build range options
    current_year = today.year
    range_options = [
        ("12", "Last 12 months"),
        ("ytd", f"Year to Date ({current_year})"),
    ]
    # Add previous years
    for year in range(current_year, current_year - 5, -1):
        range_options.append((str(year), str(year)))

    # Calculate date range based on selection
    if selected_range == "ytd":
        first_day = date(current_year, 1, 1)
        last_day = date(current_year, 12, 31)  # Include future transactions through end of year
    elif selected_range == "12":
        # Last 12 months
        first_day = date(today.year - 1, today.month, 1)
        last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])
    else:
        # Specific year
        try:
            year = int(selected_range)
            first_day = date(year, 1, 1)
            last_day = date(year, 12, 31)
        except (ValueError, TypeError):
            # Fallback to last 12 months
            first_day = date(today.year - 1, today.month, 1)
            last_day = date(today.year, today.month, monthrange(today.year, today.month)[1])

    # ---------- Derived history from Transfers + Expenses (filtered by date range) ----------

    # Transfers tagged with this bucket (within date range)
    transfer_qs = (
        Transfer.objects.filter(
            withholding_category=category,
            date__range=(first_day, last_day)
        )
        .select_related("from_account", "to_account")
        .order_by("date", "id")
    )

    # Expenses funded from this bucket (within date range)
    expense_qs = (
        Expense.objects.filter(
            withholding_category=category,
            date__range=(first_day, last_day)
        )
        .select_related("category", "bank_account", "rental_unit", "cra_category")
        .order_by("date", "id")
    )

    derived_events = []
    bucket_account = category.account

    # Transfers
    for t in transfer_qs:
        desc = t.description or "Transfer"
        signed = Decimal("0.00")
        kind = "transfer_other"

        if bucket_account:
            if t.to_account_id == bucket_account.id:
                # Money moved into the bucket
                signed = t.amount
                kind = "transfer_in"
            elif t.from_account_id == bucket_account.id:
                # Money moved out of the bucket
                signed = -t.amount
                kind = "transfer_out"

        derived_events.append(
            {
                "kind": kind,
                "date": t.date,
                "signed_amount": signed,
                "description": desc,
                "transfer_id": t.id,
                "expense_id": None,
            }
        )

    # Expenses
    for e in expense_qs:
        if e.vendor_name:
            desc = e.vendor_name
        elif e.category_id:
            desc = e.category.name
        else:
            desc = "Expense"

        # Expense funded from bucket always reduces it
        signed = -e.amount

        # Prepare expense data for JavaScript
        import json
        expense_data = {
            'id': e.id,
            'date': e.date.strftime('%Y-%m-%d'),
            'vendor_name': e.vendor_name or '',
            'category_name': e.category.name if e.category else '',
            'location': e.location or '',
            'amount': str(e.amount),
            'notes': e.notes or '',
            'rental_unit_id': e.rental_unit_id or '',
            'cra_category_id': e.cra_category_id or '',
            'rental_business_use_pct': str(e.rental_business_use_pct) if e.rental_business_use_pct else '',
            'bank_account_id': e.bank_account_id or '',
        }

        derived_events.append(
            {
                "kind": "expense",
                "date": e.date,
                "signed_amount": signed,
                "description": desc,
                "transfer_id": None,
                "expense_id": e.id,
                "expense_obj": json.dumps(expense_data),
            }
        )

    # Sort by date (ascending) to build running total
    derived_events.sort(key=lambda ev: (ev["date"],))

    derived_running = Decimal("0.00")
    derived_rows_chron = []
    for ev in derived_events:
        derived_running += ev["signed_amount"]
        row_data = {
            "date": ev["date"],
            "kind": ev["kind"],
            "description": ev["description"],
            "signed_amount": ev["signed_amount"],
            "balance_after": derived_running,
            "transfer_id": ev.get("transfer_id"),
            "expense_id": ev.get("expense_id"),
        }
        # Add expense_obj if present
        if "expense_obj" in ev:
            row_data["expense_obj"] = ev["expense_obj"]
        derived_rows_chron.append(row_data)

    # Calculate total balance (all-time, not just the selected range)
    all_transfer_qs = Transfer.objects.filter(withholding_category=category).select_related("from_account", "to_account")
    all_expense_qs = Expense.objects.filter(withholding_category=category)

    derived_balance = Decimal("0.00")
    bucket_account = category.account

    for t in all_transfer_qs:
        if bucket_account:
            if t.to_account_id == bucket_account.id:
                derived_balance += t.amount
            elif t.from_account_id == bucket_account.id:
                derived_balance -= t.amount

    for e in all_expense_qs:
        derived_balance -= e.amount

    # Calculate total for the selected range
    range_total = sum(ev["signed_amount"] for ev in derived_events)

    # Context data for modals
    all_categories = Category.objects.filter(is_archived=False).order_by("name")
    accounts = BankAccount.objects.all().order_by("name")
    all_rental_units = RentalUnit.objects.select_related("property").order_by("property__name", "name")
    cra_categories = CRARentalExpenseCategory.objects.all().order_by("name")
    withholding_categories = WithholdingCategory.objects.select_related("account").all().order_by("account__name", "name")
    income_categories = IncomeCategory.objects.all().order_by("name")

    context = {
        "category": category,
        "derived_rows": list(reversed(derived_rows_chron)),  # newest first for display
        "derived_balance": derived_balance,  # Current balance (all-time)
        "range_total": range_total,  # Total change in selected period
        "range_options": range_options,
        "selected_range": selected_range,
        # For modals
        "all_categories": all_categories,
        "accounts": accounts,
        "all_rental_units": all_rental_units,
        "cra_categories": cra_categories,
        "withholding_categories": withholding_categories,
        "income_categories": income_categories,
    }
    return render(request, "withholding_category_detail.html", context)


@require_POST
def update_withholding_transaction(request, pk):
    tx = get_object_or_404(WithholdingTransaction, pk=pk)
    category_pk = tx.category.pk

    if request.POST.get("delete_withholding") == "1":
        tx.delete()
        return redirect("withholding_category_detail", pk=category_pk)

    date_str = request.POST.get("date")
    amount_str = request.POST.get("amount")
    note = request.POST.get("note", "")

    if date_str:
        try:
            tx.date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    if amount_str:
        cleaned = amount_str.replace("$", "").replace(",", "").strip()
        try:
            tx.amount = Decimal(cleaned)
        except InvalidOperation:
            pass

    tx.note = note
    tx.save()

    return redirect("withholding_category_detail", pk=category_pk)

@require_http_methods(["GET", "POST"])
def expense_edit(request, expense_id):
    expense = get_object_or_404(
        Expense.objects.select_related("category", "bank_account", "rental_unit", "cra_category"),
        pk=expense_id
    )

    if request.method == "POST":
        # Delete attachment
        delete_attachment_id = request.POST.get("delete_attachment_id")
        if delete_attachment_id:
            att = get_object_or_404(ExpenseAttachment, pk=delete_attachment_id, expense=expense)
            if att.file:
                att.file.delete(save=False)
            att.delete()
            messages.success(request, "Attachment deleted.")
            return redirect("expense_edit", expense_id=expense.id)

        # IMPORTANT: do uploads directly from request.FILES
        files = request.FILES.getlist("files")
        print("FILES KEYS:", list(request.FILES.keys()))
        print("FILES COUNT:", len(files))

        # Check if this is a receipt-only upload (from tax summary modal)
        # If only expense_id and files are present (no other expense form fields), handle separately
        is_receipt_only_upload = (
            request.POST.get("expense_id") and
            files and
            not any(field in request.POST for field in ["date", "amount", "vendor_name"])
        )

        if is_receipt_only_upload:
            # Handle receipt-only upload without form validation
            created = 0
            for f in files:
                if not f:
                    continue
                ExpenseAttachment.objects.create(
                    expense=expense,
                    file=f,
                    original_name=getattr(f, "name", "") or "",
                )
                created += 1

            if created:
                messages.success(request, f"Uploaded {created} receipt(s) successfully.")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # AJAX request from modal
                    from django.http import JsonResponse
                    return JsonResponse({"success": True, "message": f"Uploaded {created} receipt(s)"})
                return redirect("expense_edit", expense_id=expense.id)
            else:
                messages.error(request, "No files were uploaded.")
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    from django.http import JsonResponse
                    return JsonResponse({"success": False, "error": "No files uploaded"}, status=400)
                return redirect("expense_edit", expense_id=expense.id)

        # Otherwise, process as normal expense edit with form validation
        form = ExpenseEditForm(request.POST, instance=expense)

        if form.is_valid():
            form.save()

            created = 0
            for f in files:
                # skip empty placeholders (some browsers can submit empties)
                if not f:
                    continue
                ExpenseAttachment.objects.create(
                    expense=expense,
                    file=f,
                    original_name=getattr(f, "name", "") or "",
                )
                created += 1

            if created:
                messages.success(request, f"Saved expense and uploaded {created} attachment(s).")
            else:
                messages.success(request, "Saved expense (no attachments uploaded).")

            return redirect("expense_edit", expense_id=expense.id)

        # If the expense form is invalid, show errors (and keep user on page)
        messages.error(request, "Please correct the errors below.")
    else:
        form = ExpenseEditForm(instance=expense)

    attachments = expense.attachments.all()

    return render(request, "expense_edit.html", {
        "expense": expense,
        "form": form,
        "attachments": attachments,
    })


@require_http_methods(["GET", "POST"])
def income_edit(request, income_id):
    income = get_object_or_404(
        Income.objects.select_related("category", "income_category", "rental_unit", "bank_account"),
        pk=income_id
    )

    if request.method == "POST":
        form = IncomeEditForm(request.POST, instance=income)

        if form.is_valid():
            form.save()
            messages.success(request, "Income updated successfully.")
            # Redirect back to dashboard with the income's month
            return redirect(f"{reverse('dashboard')}?month={income.date.year}-{income.date.month:02d}")

        # If the form is invalid, show errors (and keep user on page)
        messages.error(request, "Please correct the errors below.")
    else:
        form = IncomeEditForm(instance=income)

    return render(request, "income_edit.html", {
        "income": income,
        "form": form,
    })


@require_http_methods(["GET", "POST"])
def transfer_edit(request, transfer_id):
    transfer = get_object_or_404(
        Transfer.objects.select_related("from_account", "to_account", "withholding_category"),
        pk=transfer_id
    )

    if request.method == "POST":
        form = TransferEditForm(request.POST, instance=transfer)

        if form.is_valid():
            form.save()
            messages.success(request, "Transfer updated successfully.")
            # Redirect back to dashboard with the transfer's month
            return redirect(f"{reverse('dashboard')}?month={transfer.date.year}-{transfer.date.month:02d}")

        # If the form is invalid, show errors (and keep user on page)
        messages.error(request, "Please correct the errors below.")
    else:
        form = TransferEditForm(instance=transfer)

    return render(request, "transfer_edit.html", {
        "transfer": transfer,
        "form": form,
    })


@require_http_methods(["GET", "POST"])
def balance_adjustment_edit(request, adjustment_id):
    adjustment = get_object_or_404(
        BalanceAdjustment.objects.select_related("bank_account"),
        pk=adjustment_id
    )

    if request.method == "POST":
        # Handle delete action
        if request.POST.get("action") == "delete":
            account_id = adjustment.bank_account_id
            adjustment.delete()
            messages.success(request, "Balance adjustment deleted successfully.")
            return redirect('bank_account_detail', account_id=account_id)

        # Handle edit action
        form = BalanceAdjustmentEditForm(request.POST, instance=adjustment)

        if form.is_valid():
            form.save()
            messages.success(request, "Balance adjustment updated successfully.")
            return redirect('bank_account_detail', account_id=adjustment.bank_account_id)

        # If the form is invalid, show errors
        messages.error(request, "Please correct the errors below.")
    else:
        form = BalanceAdjustmentEditForm(instance=adjustment)

    return render(request, "balance_adjustment_edit.html", {
        "adjustment": adjustment,
        "form": form,
    })


@require_http_methods(['GET'])
def get_transfer_api(request, transfer_id):
    """API endpoint to get transfer data as JSON."""
    try:
        transfer = Transfer.objects.select_related(
            'from_account', 'to_account', 'withholding_category', 'parent_transfer'
        ).prefetch_related('splits').get(id=transfer_id)

        data = {
            'id': transfer.id,
            'date': transfer.date.isoformat(),
            'amount': str(transfer.amount),
            'description': transfer.description,
            'notes': transfer.notes,
            'from_account_id': transfer.from_account_id,
            'to_account_id': transfer.to_account_id,
            'withholding_category_id': transfer.withholding_category_id,
            'is_split_parent': transfer.is_split_parent,
            'parent_transfer_id': transfer.parent_transfer_id,
            'split_count': transfer.split_count,
        }

        if transfer.is_split_parent:
            data['splits'] = [{
                'id': s.id,
                'amount': str(s.amount),
                'from_account_id': s.from_account_id,
                'to_account_id': s.to_account_id,
                'withholding_category_id': s.withholding_category_id,
                'notes': s.notes,
                'split_order': s.split_order,
            } for s in transfer.splits.order_by('split_order')]

        return JsonResponse(data)
    except Transfer.DoesNotExist:
        return JsonResponse({'error': 'Not found'}, status=404)


@transaction.atomic
def handle_transfer_edit(request, is_ajax=False):
    """Handle transfer editing including splits."""
    transfer_id = request.POST.get('transfer_id')
    transfer = get_object_or_404(Transfer, pk=transfer_id)

    # Handle delete
    if 'delete_transfer' in request.POST:
        transfer_id_save = transfer.id
        transfer.delete()  # CASCADE deletes children automatically

        if is_ajax:
            return JsonResponse({
                'deleted': True,
                'transaction_id': transfer_id_save
            })

        return redirect(f"/?month={request.POST.get('month', '')}")

    # Handle split mode
    if request.POST.get('split_mode') == '1':
        return handle_transfer_split(request, transfer, is_ajax)

    # Handle normal update (non-split or split child edit)
    return handle_transfer_update(request, transfer, is_ajax)


@transaction.atomic
def handle_transfer_split(request, transfer, is_ajax=False):
    """Create or update splits for a transfer."""

    # Collect split data from POST
    splits = []
    i = 0
    while f'split-{i}-amount' in request.POST:
        try:
            amount = Decimal(request.POST.get(f'split-{i}-amount', '0'))
            from_account_id = request.POST.get(f'split-{i}-from_account') or None
            to_account_id = request.POST.get(f'split-{i}-to_account') or None
            withholding_id = request.POST.get(f'split-{i}-withholding_category') or None
            notes = request.POST.get(f'split-{i}-notes', '')

            if amount > 0:
                splits.append({
                    'amount': amount,
                    'from_account_id': from_account_id,
                    'to_account_id': to_account_id,
                    'withholding_category_id': withholding_id,
                    'notes': notes,
                })
        except (InvalidOperation, ValueError):
            continue
        i += 1

    # Validate sum
    total = sum(s['amount'] for s in splits)
    tolerance = Decimal('0.005')
    if abs(transfer.amount - total) > tolerance:
        if is_ajax:
            return JsonResponse({
                'error': 'Split amounts do not match transfer total'
            }, status=400)
        return redirect(f"/?month={request.POST.get('month', '')}")

    # Mark as split parent and delete old splits
    transfer.is_split_parent = True
    transfer.save()
    transfer.splits.all().delete()

    # Create new splits
    for idx, split_data in enumerate(splits):
        Transfer.objects.create(
            date=transfer.date,
            amount=split_data['amount'],
            description=transfer.description,
            notes=split_data['notes'],
            from_account_id=split_data['from_account_id'],
            to_account_id=split_data['to_account_id'],
            withholding_category_id=split_data['withholding_category_id'],
            parent_transfer=transfer,
            split_order=idx + 1,
        )

    if is_ajax:
        return JsonResponse({
            'success': True,
            'message': 'Transfer split successfully'
        })

    return redirect(f"/?month={request.POST.get('month', '')}")


def handle_transfer_update(request, transfer, is_ajax=False):
    """Update a single transfer (non-split or split child)."""
    try:
        # Date
        transfer.date = datetime.strptime(request.POST["date"], "%Y-%m-%d").date()

        # Amount
        amount_str = (request.POST.get("amount") or "").strip()
        if amount_str:
            try:
                transfer.amount = Decimal(amount_str)
            except (InvalidOperation, ValueError):
                if is_ajax:
                    return JsonResponse({
                        'error': 'Invalid amount format'
                    }, status=400)
                pass

        # From / To accounts
        from_account_id = (request.POST.get("from_account") or "").strip()
        to_account_id = (request.POST.get("to_account") or "").strip()

        transfer.from_account = (
            get_object_or_404(BankAccount, pk=from_account_id)
            if from_account_id else None
        )
        transfer.to_account = (
            get_object_or_404(BankAccount, pk=to_account_id)
            if to_account_id else None
        )

        # Withholding bucket
        bucket_id = (request.POST.get("withholding_category") or "").strip()
        if bucket_id:
            transfer.withholding_category = get_object_or_404(
                WithholdingCategory,
                pk=bucket_id
            )
        else:
            transfer.withholding_category = None

        # Notes
        transfer.notes = request.POST.get("notes", "")

        transfer.save()

        # If this is a split child, validate parent still sums correctly
        if transfer.parent_transfer and not transfer.parent_transfer.validate_split_amounts():
            # Log warning or notify user
            pass

        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': 'Transfer updated successfully'
            })

        if transfer.date:
            selected_month_param = f"{transfer.date.year:04d}-{transfer.date.month:02d}"
        else:
            selected_month_param = request.POST.get('month', '')

        return redirect(f"/?month={selected_month_param}")
    except Exception as e:
        if is_ajax:
            return JsonResponse({
                'error': f'Error updating transfer: {str(e)}'
            }, status=400)
        raise


# ============================================
# MONTH-END CLOSE WIZARD
# ============================================

def create_comprehensive_backup(month_str, description="Month-end close"):
    """
    Create a comprehensive backup including:
    1. JSON data export
    2. SQLite database file
    3. Media files (receipts/attachments)
    4. Zip everything together

    Returns: (zip_filename, backup_info_dict)
    """
    import os
    import shutil
    import zipfile
    from pathlib import Path
    from django.core.management import call_command
    from django.conf import settings
    from datetime import datetime

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    # Temporary directory for backup components
    temp_backup_dir = os.path.join(backup_dir, f'temp_{timestamp}')
    os.makedirs(temp_backup_dir, exist_ok=True)

    backup_info = {
        'timestamp': timestamp,
        'month': month_str,
        'description': description,
        'components': []
    }

    try:
        # 1. JSON Data Export
        print(f"[BACKUP] Step 1/5: Creating JSON data export...")
        json_filename = f'data_{month_str}_{timestamp}.json'
        json_path = os.path.join(temp_backup_dir, json_filename)

        with open(json_path, 'w') as f:
            call_command('dumpdata', 'home', indent=2, stdout=f)

        json_size = os.path.getsize(json_path)
        print(f"[BACKUP] JSON export complete: {json_size/1024:.1f} KB")
        backup_info['components'].append({
            'name': 'JSON Data Export',
            'filename': json_filename,
            'size': json_size
        })

        # 2. SQLite Database File
        print(f"[BACKUP] Step 2/5: Copying database file...")
        db_path = os.path.join(settings.BASE_DIR, 'db.sqlite3')
        if os.path.exists(db_path):
            db_filename = f'database_{month_str}_{timestamp}.sqlite3'
            db_backup_path = os.path.join(temp_backup_dir, db_filename)
            shutil.copy2(db_path, db_backup_path)

            db_size = os.path.getsize(db_backup_path)
            print(f"[BACKUP] Database copied: {db_size/1024/1024:.1f} MB")
            backup_info['components'].append({
                'name': 'SQLite Database',
                'filename': db_filename,
                'size': db_size
            })

        # 3. Media Files (receipts/attachments)
        print(f"[BACKUP] Step 3/5: Copying media files (receipts)...")
        media_root = settings.MEDIA_ROOT
        if os.path.exists(media_root) and os.listdir(media_root):
            media_backup_dir = os.path.join(temp_backup_dir, 'media')
            shutil.copytree(media_root, media_backup_dir)
            print(f"[BACKUP] Media files copied to temp directory")

            # Count media files
            print(f"[BACKUP] Calculating media file statistics...")
            media_files = sum(1 for _ in Path(media_backup_dir).rglob('*') if _.is_file())
            media_size = sum(f.stat().st_size for f in Path(media_backup_dir).rglob('*') if f.is_file())
            print(f"[BACKUP] Media files: {media_files} files, {media_size/1024/1024:.1f} MB")

            backup_info['components'].append({
                'name': 'Media Files (Receipts)',
                'filename': 'media/',
                'size': media_size,
                'count': media_files
            })

        # 4. Create README
        print(f"[BACKUP] Step 4/5: Creating README file...")
        readme_path = os.path.join(temp_backup_dir, 'README.txt')
        with open(readme_path, 'w') as f:
            f.write(f"FinchFinance Backup - {month_str}\n")
            f.write(f"Created: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Description: {description}\n")
            f.write(f"\n")
            f.write(f"CONTENTS:\n")
            f.write(f"-" * 60 + "\n")
            for comp in backup_info['components']:
                f.write(f"- {comp['name']}: {comp['filename']}\n")
                f.write(f"  Size: {comp['size']:,} bytes ({comp['size']/1024/1024:.2f} MB)\n")
                if 'count' in comp:
                    f.write(f"  Files: {comp['count']}\n")
            f.write(f"\n")
            f.write(f"RESTORE INSTRUCTIONS:\n")
            f.write(f"-" * 60 + "\n")
            f.write(f"1. Extract this zip file\n")
            f.write(f"2. Quick restore (database file):\n")
            f.write(f"   - Copy {db_filename if 'db_filename' in locals() else 'database_*.sqlite3'} to project root as db.sqlite3\n")
            f.write(f"   - Copy media/ folder to project root\n")
            f.write(f"3. OR Data-only restore (JSON):\n")
            f.write(f"   - python manage.py migrate\n")
            f.write(f"   - python manage.py loaddata {json_filename}\n")
            f.write(f"   - Copy media/ folder to project root\n")

        # 5. Create ZIP file
        print(f"[BACKUP] Step 5/5: Creating ZIP archive...")
        zip_filename = f'monthend_{month_str}_{timestamp}.zip'
        zip_path = os.path.join(backup_dir, zip_filename)

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            file_count = 0
            for root, dirs, files in os.walk(temp_backup_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_backup_dir)
                    zipf.write(file_path, arcname)
                    file_count += 1
                    if file_count % 10 == 0:
                        print(f"[BACKUP] Compressed {file_count} files...")

        print(f"[BACKUP] ZIP created: {zip_filename}")

        # Clean up temp directory
        print(f"[BACKUP] Cleaning up temporary files...")
        shutil.rmtree(temp_backup_dir)

        # Final backup info
        backup_info['zip_filename'] = zip_filename
        backup_info['zip_size'] = os.path.getsize(zip_path)
        backup_info['zip_path'] = zip_path

        print(f"[BACKUP] ✅ Backup complete: {backup_info['zip_size']/1024/1024:.1f} MB")
        print(f"[BACKUP] Location: {zip_path}")

        return zip_filename, backup_info

    except Exception as e:
        # Clean up on error
        if os.path.exists(temp_backup_dir):
            shutil.rmtree(temp_backup_dir)
        raise e


def month_end_wizard(request):
    """
    Multi-step wizard for closing a month:
    Step 1: Select month
    Step 2: Review financial summary
    Step 3: Review account snapshots
    Step 4: Review/edit net worth
    Step 5: Confirm and execute
    """
    import os
    import json
    from django.core.management import call_command
    from .models import MonthEndClose, AccountSnapshot, NetWorthSnapshot, RentalProperty

    step = request.GET.get('step', '1')
    month_str = request.GET.get('month', '')

    # Step 1: Select Month
    if step == '1':
        # Get months with transactions
        all_expenses = Expense.objects.all().values_list('date', flat=True)
        all_income = Income.objects.all().values_list('date', flat=True)
        all_transfers = Transfer.objects.all().values_list('date', flat=True)

        all_dates = list(all_expenses) + list(all_income) + list(all_transfers)

        months_with_data = {}
        for d in all_dates:
            month_key = d.strftime('%Y-%m')
            month_first = date(d.year, d.month, 1)

            if month_first not in months_with_data:
                # Check if already closed
                is_closed = MonthEndClose.objects.filter(month=month_first, is_locked=True).exists()
                months_with_data[month_first] = {
                    'key': month_key,
                    'display': d.strftime('%B %Y'),
                    'is_closed': is_closed,
                    'can_close': month_first < date.today() and not is_closed
                }

        context = {
            'months': sorted(months_with_data.items(), reverse=True)[:12],  # Last 12 months
            'step': 1,
        }
        return render(request, 'month_end_wizard.html', context)

    # Parse selected month
    try:
        year, month = map(int, month_str.split('-'))
        month_first_day = date(year, month, 1)
        _, last_day_num = monthrange(year, month)
        month_last_day = date(year, month, last_day_num)
        month_display = month_first_day.strftime('%B %Y')
    except:
        messages.error(request, 'Invalid month selected')
        return redirect('month_end_wizard')

    # Check if already closed
    existing_close = MonthEndClose.objects.filter(month=month_first_day).first()
    if existing_close and existing_close.is_locked and step != '1':
        messages.warning(request, f'{month_display} is already closed')
        return redirect('month_end_wizard')

    # Calculate data for all steps
    income_entries = Income.objects.filter(date__range=(month_first_day, month_last_day))
    expense_entries = Expense.objects.filter(date__range=(month_first_day, month_last_day))
    transfer_entries = Transfer.objects.filter(date__range=(month_first_day, month_last_day))

    # Exclude Business Reimbursement from income totals (similar to excluding Business Expense)
    total_income = sum(
        i.amount for i in income_entries
        if not i.income_category or i.income_category.name != 'Business Reimbursement'
    )
    total_expenses = sum(e.amount for e in expense_entries)
    total_transfers = sum(t.amount for t in transfer_entries)
    net_savings = total_income - total_expenses
    transaction_count = income_entries.count() + expense_entries.count() + transfer_entries.count()

    # Step 2 POST: Save excess and create/update bucket
    if step == '2' and request.method == 'POST':
        if request.POST.get('save_excess') == '1':
            excess_amount = Decimal(request.POST.get('excess_amount', '0'))
            excess_account_id = request.POST.get('excess_account')

            if excess_amount > 0 and excess_account_id:
                try:
                    account = BankAccount.objects.get(
                        id=excess_account_id,
                        is_withholding_account=True
                    )

                    # Get or create "Excess/Surplus" bucket
                    bucket, created = WithholdingCategory.objects.get_or_create(
                        account=account,
                        name="Excess/Surplus",
                        defaults={
                            'monthly_target': None,
                            'target_amount': Decimal('0.00'),
                        }
                    )

                    # Create transfer INTO the excess bucket
                    Transfer.objects.create(
                        date=month_last_day,
                        amount=excess_amount,
                        description=f"Month-end excess savings for {month_display}",
                        to_account=account,
                        withholding_category=bucket,
                        notes=f"Automatic excess savings from month-end close"
                    )

                    messages.success(
                        request,
                        f"Saved ${excess_amount} excess to {bucket.name} bucket in {account.name}"
                    )

                except BankAccount.DoesNotExist:
                    messages.error(request, "Selected withholding account not found.")
                    return redirect(f"month_end_wizard?step=2&month={month_str}")

            # Continue to Step 3
            return redirect(f"month_end_wizard?step=3&month={month_str}")

    # Step 2 GET: Review Enhanced Financial Summary
    if step == '2':
        # Income breakdown by category (exclude Business Reimbursement)
        income_breakdown = []
        business_reimbursement_total = Decimal('0.00')

        for inc_category in IncomeCategory.objects.all():
            amount = income_entries.filter(income_category=inc_category).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')

            if amount > 0:
                if inc_category.name == 'Business Reimbursement':
                    business_reimbursement_total = amount
                else:
                    income_breakdown.append({
                        'name': inc_category.name,
                        'amount': amount
                    })

        # Planned expense analysis (categories with monthly_limit > 0, exclude Business Expense)
        planned_categories = Category.objects.filter(is_archived=False, monthly_limit__gt=0).exclude(name='Business Expense')
        planned_breakdown = []
        total_planned_budget = Decimal('0.00')
        total_planned_spent = Decimal('0.00')

        for category in planned_categories:
            spent = expense_entries.filter(category=category).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')

            budget = category.monthly_limit
            variance = spent - budget

            planned_breakdown.append({
                'name': category.name,
                'budget': budget,
                'spent': spent,
                'variance': variance,
                'status': 'over' if variance > 0 else 'under' if variance < 0 else 'on_target'
            })

            total_planned_budget += budget
            total_planned_spent += spent

        planned_variance = total_planned_spent - total_planned_budget

        # Unplanned expense analysis (categories with monthly_limit = 0 or NULL, exclude Business Expense)
        from django.db.models import Q
        unplanned_categories = Category.objects.filter(
            is_archived=False
        ).filter(
            Q(monthly_limit__isnull=True) | Q(monthly_limit=0)
        ).exclude(name='Business Expense')
        unplanned_breakdown = []
        total_unplanned_spent = Decimal('0.00')

        for category in unplanned_categories:
            spent = expense_entries.filter(category=category).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')

            if spent > 0:  # Only include if there was spending
                unplanned_breakdown.append({
                    'name': category.name,
                    'spent': spent
                })
                total_unplanned_spent += spent

        # Business Expense tracking (excluded from net calculations but shown for reference)
        business_expense_cat = Category.objects.filter(name='Business Expense').first()
        business_expense_total = Decimal('0.00')
        if business_expense_cat:
            business_expense_total = expense_entries.filter(category=business_expense_cat).aggregate(
                total=Sum('amount')
            )['total'] or Decimal('0.00')

        # Withholding/Savings analysis (buckets with monthly_target)
        withholding_buckets = WithholdingCategory.objects.exclude(monthly_target__isnull=True)
        withholding_breakdown = []
        total_withholding_target = Decimal('0.00')
        total_withholding_actual = Decimal('0.00')

        for bucket in withholding_buckets:
            # Contributions this month (transfers INTO the bucket's account)
            contributions = transfer_entries.filter(
                withholding_category=bucket,
                to_account=bucket.account
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')

            target = bucket.monthly_target
            variance = contributions - target

            withholding_breakdown.append({
                'name': bucket.name,
                'target': target,
                'actual': contributions,
                'variance': variance,
                'status': 'over' if variance > 0 else 'under' if variance < 0 else 'on_target'
            })

            total_withholding_target += target
            total_withholding_actual += contributions

        withholding_variance = total_withholding_actual - total_withholding_target

        # Excess calculation (excludes Business Expense, includes withholding contributions)
        # Calculate true surplus: Income - (Planned + Unplanned + Withholding)
        true_surplus = total_income - total_planned_spent - total_unplanned_spent - total_withholding_actual

        # Keep net_savings for overall display (includes Business Expense)
        net_surplus = total_income - total_expenses

        if true_surplus > 0:
            # Only calculate excess if we have actual surplus after all allocations
            planned_under_spend = sum(
                (item['budget'] - item['spent'])
                for item in planned_breakdown
                if item['variance'] < 0
            )
            excess_to_save = true_surplus
            deficit_amount = Decimal('0.00')
        else:
            planned_under_spend = Decimal('0.00')
            excess_to_save = Decimal('0.00')
            deficit_amount = abs(true_surplus) if true_surplus < 0 else Decimal('0.00')

        # Get withholding accounts for excess bucket selection
        withholding_accounts = BankAccount.objects.filter(
            is_withholding_account=True,
            is_active=True
        ).order_by('name')

        context = {
            'step': 2,
            'month': month_str,
            'month_display': month_display,

            # Existing totals
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_savings': net_savings,
            'total_transfers': total_transfers,
            'transaction_count': transaction_count,
            'income_count': income_entries.count(),
            'expense_count': expense_entries.count(),
            'transfer_count': transfer_entries.count(),

            # NEW: Income breakdown
            'income_breakdown': income_breakdown,
            'business_reimbursement_total': business_reimbursement_total,

            # NEW: Planned expense analysis
            'planned_breakdown': planned_breakdown,
            'total_planned_budget': total_planned_budget,
            'total_planned_spent': total_planned_spent,
            'planned_variance': planned_variance,

            # NEW: Unplanned expense analysis
            'unplanned_breakdown': unplanned_breakdown,
            'total_unplanned_spent': total_unplanned_spent,

            # NEW: Business Expense (excluded from net but shown for reference)
            'business_expense_total': business_expense_total,

            # NEW: Withholding analysis
            'withholding_breakdown': withholding_breakdown,
            'total_withholding_target': total_withholding_target,
            'total_withholding_actual': total_withholding_actual,
            'withholding_variance': withholding_variance,

            # NEW: Excess savings
            'net_surplus': net_surplus,
            'true_surplus': true_surplus,
            'has_surplus': true_surplus > 0,
            'excess_to_save': excess_to_save,
            'planned_under_spend': planned_under_spend,
            'deficit_amount': deficit_amount,
            'withholding_accounts': withholding_accounts,
        }
        return render(request, 'month_end_wizard.html', context)

    # Step 3: Review Account Snapshots
    elif step == '3':
        # Get previous month's close for comparison
        prev_close = MonthEndClose.objects.filter(
            month__lt=month_first_day
        ).order_by('-month').first()

        regular_accounts = []
        retirement_accounts = []
        total_regular_balance = Decimal('0')
        total_retirement_balance = Decimal('0')

        for account in BankAccount.objects.filter(is_active=True).order_by('institution', 'name'):
            balance = account.current_balance or Decimal('0')

            # Get previous balance if available
            prev_balance = Decimal('0')
            increase = Decimal('0')
            if prev_close:
                prev_snapshot = AccountSnapshot.objects.filter(
                    month_close=prev_close,
                    bank_account=account
                ).first()
                if prev_snapshot:
                    prev_balance = prev_snapshot.balance
                    increase = balance - prev_balance

            account_data = {
                'id': account.id,
                'name': f'{account.institution} - {account.name}',
                'balance': balance,
                'prev_balance': prev_balance,
                'increase': increase,
            }

            # Separate retirement accounts
            if account.account_type == 'RETIREMENT':
                retirement_accounts.append(account_data)
                total_retirement_balance += balance
            else:
                regular_accounts.append(account_data)
                total_regular_balance += balance

        # Calculate total retirement increase
        prev_retirement_total = Decimal('0')
        if prev_close:
            for acc in retirement_accounts:
                prev_retirement_total += acc['prev_balance']
        retirement_increase = total_retirement_balance - prev_retirement_total

        context = {
            'step': 3,
            'month': month_str,
            'month_display': month_display,
            'regular_accounts': regular_accounts,
            'retirement_accounts': retirement_accounts,
            'total_regular_balance': total_regular_balance,
            'total_retirement_balance': total_retirement_balance,
            'retirement_increase': retirement_increase,
            'prev_close_month': prev_close.month_display if prev_close else None,
        }
        return render(request, 'month_end_wizard.html', context)

    # Step 4: Review/Edit Net Worth
    elif step == '4':
        # Calculate net worth components
        # Separate liquid assets from investment assets (TFSA, RETIREMENT)
        liquid_assets = Decimal('0')
        investment_assets = Decimal('0')

        for acc in BankAccount.objects.filter(is_active=True):
            balance = acc.current_balance or Decimal('0')
            if acc.account_type in ['TFSA', 'RETIREMENT']:
                investment_assets += balance
            else:
                liquid_assets += balance

        # Property equity - check if before or after Foxview purchase
        property_equity = Decimal('0')
        property_notes = []

        # Arnprior (always include if exists)
        arnprior = RentalProperty.objects.filter(name__icontains='Arnprior', is_active=True).first()
        if arnprior and arnprior.equity:
            property_equity += arnprior.equity
            property_notes.append(f'Arnprior: ${arnprior.equity:,.2f}')

        # Foxview (only if Feb 2026 or later)
        if month_first_day >= date(2026, 2, 1):
            foxview = RentalProperty.objects.filter(name__icontains='Foxview', is_active=True).first()
            if foxview and foxview.equity:
                property_equity += foxview.equity
                property_notes.append(f'Foxview: ${foxview.equity:,.2f}')
        else:
            property_notes.append('Foxview: N/A (purchased Feb 1, 2026)')

        total_net_worth = liquid_assets + investment_assets + property_equity

        context = {
            'step': 4,
            'month': month_str,
            'month_display': month_display,
            'liquid_assets': liquid_assets,
            'investment_assets': investment_assets,
            'property_value': property_equity,
            'property_notes': property_notes,
            'liabilities': Decimal('0'),
            'total_net_worth': total_net_worth,
        }
        return render(request, 'month_end_wizard.html', context)

    # Step 5: Confirm and Execute
    elif step == '5' and request.method == 'POST':
        print(f"\n{'='*60}")
        print(f"[MONTH-END] Starting month-end close for {month_str}")
        print(f"{'='*60}\n")
        try:
            # Get all transactions for the month
            print(f"[MONTH-END] Loading transaction data...")
            income_entries = Income.objects.filter(date__range=(month_first_day, month_last_day))
            expense_entries = Expense.objects.filter(date__range=(month_first_day, month_last_day))
            transfer_entries = Transfer.objects.filter(date__range=(month_first_day, month_last_day))

            # Calculate totals (excluding Business Reimbursement from income)
            total_income = sum(
                i.amount for i in income_entries
                if not i.income_category or i.income_category.name != 'Business Reimbursement'
            )
            total_expenses = sum(e.amount for e in expense_entries)
            net_savings = total_income - total_expenses
            transaction_count = len(income_entries) + len(expense_entries) + len(transfer_entries)
            total_transfers = sum(t.amount for t in transfer_entries)
            print(f"[MONTH-END] Processed {transaction_count} transactions")

            # Recalculate enhanced summary data for storage
            print(f"[MONTH-END] Calculating enhanced summary data...")
            # Planned expense totals (monthly_limit > 0, exclude Business Expense)
            from django.db.models import Q
            planned_categories = Category.objects.filter(is_archived=False, monthly_limit__gt=0).exclude(name='Business Expense')
            total_planned_budget = Decimal('0.00')
            total_planned_spent = Decimal('0.00')
            for category in planned_categories:
                total_planned_budget += category.monthly_limit
                spent = expense_entries.filter(category=category).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0.00')
                total_planned_spent += spent

            # Unplanned expense totals (monthly_limit = 0 or NULL, exclude Business Expense)
            unplanned_categories = Category.objects.filter(
                is_archived=False
            ).filter(
                Q(monthly_limit__isnull=True) | Q(monthly_limit=0)
            ).exclude(name='Business Expense')
            total_unplanned_spent = Decimal('0.00')
            for category in unplanned_categories:
                spent = expense_entries.filter(category=category).aggregate(
                    total=Sum('amount')
                )['total'] or Decimal('0.00')
                total_unplanned_spent += spent

            # Withholding totals
            withholding_buckets = WithholdingCategory.objects.exclude(monthly_target__isnull=True)
            total_withholding_target = Decimal('0.00')
            total_withholding_actual = Decimal('0.00')
            for bucket in withholding_buckets:
                total_withholding_target += bucket.monthly_target
                contributions = transfer_entries.filter(
                    withholding_category=bucket,
                    to_account=bucket.account
                ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
                total_withholding_actual += contributions

            # Excess saved (check if an excess transfer was created)
            net_surplus = total_income - total_expenses
            excess_saved = Decimal('0.00')
            if net_surplus > 0:
                # Look for excess transfer created during Step 2
                excess_bucket = WithholdingCategory.objects.filter(name="Excess/Surplus").first()
                if excess_bucket:
                    excess_transfer = transfer_entries.filter(
                        withholding_category=excess_bucket,
                        description__icontains="Month-end excess savings"
                    ).first()
                    if excess_transfer:
                        excess_saved = excess_transfer.amount

            # Create comprehensive backup (JSON + DB + Media files, all zipped)
            print(f"[MONTH-END] Starting comprehensive backup...")
            backup_filename, backup_info = create_comprehensive_backup(
                month_str,
                description=f"Month-end close for {month_display}"
            )

            # Create month-end close with all snapshots
            print(f"[MONTH-END] Creating month-end close record and snapshots...")
            with transaction.atomic():
                month_close = MonthEndClose.objects.create(
                    month=month_first_day,
                    closed_by=request.user.username if request.user.is_authenticated else 'System',
                    backup_file=backup_filename,
                    total_income=total_income,
                    total_expenses=total_expenses,
                    net_savings=net_savings,
                    total_transfers=total_transfers,
                    transaction_count=transaction_count,
                    is_locked=True,

                    # Enhanced summary fields
                    planned_budget_total=total_planned_budget,
                    planned_spent_total=total_planned_spent,
                    unplanned_spent_total=total_unplanned_spent,
                    withholding_target_total=total_withholding_target,
                    withholding_actual_total=total_withholding_actual,
                    excess_saved=excess_saved,
                )

                # Account snapshots
                for account in BankAccount.objects.filter(is_active=True):
                    balance = account.current_balance or Decimal('0')
                    if balance != 0:
                        AccountSnapshot.objects.create(
                            month_close=month_close,
                            bank_account=account,
                            balance=balance,
                        )

                # Net worth snapshot
                # Separate liquid assets from investment assets (TFSA, RETIREMENT)
                liquid_assets = Decimal('0')
                investment_assets = Decimal('0')

                for acc in BankAccount.objects.filter(is_active=True):
                    balance = acc.current_balance or Decimal('0')
                    if acc.account_type in ['TFSA', 'RETIREMENT']:
                        investment_assets += balance
                    else:
                        liquid_assets += balance

                property_equity = Decimal('0')
                property_notes_list = []

                arnprior = RentalProperty.objects.filter(name__icontains='Arnprior', is_active=True).first()
                if arnprior and arnprior.equity:
                    property_equity += arnprior.equity
                    property_notes_list.append(f'Arnprior: ${arnprior.equity:,.2f}')

                if month_first_day >= date(2026, 2, 1):
                    foxview = RentalProperty.objects.filter(name__icontains='Foxview', is_active=True).first()
                    if foxview and foxview.equity:
                        property_equity += foxview.equity
                        property_notes_list.append(f'Foxview: ${foxview.equity:,.2f}')
                else:
                    property_notes_list.append('Foxview excluded (purchased Feb 1, 2026)')

                NetWorthSnapshot.objects.create(
                    month_close=month_close,
                    total_net_worth=liquid_assets + investment_assets + property_equity,
                    liquid_assets=liquid_assets,
                    investment_assets=investment_assets,
                    property_value=property_equity,
                    liabilities=Decimal('0'),
                    notes='; '.join(property_notes_list),
                )

                # Expense category snapshots
                print(f"[MONTH-END] Creating category snapshots...")
                for category in planned_categories:
                    spent = expense_entries.filter(category=category).aggregate(
                        total=Sum('amount')
                    )['total'] or Decimal('0')
                    MonthEndExpenseCategorySnapshot.objects.create(
                        month_close=month_close,
                        category=category,
                        monthly_limit=category.monthly_limit,
                        actual_spent=spent,
                    )

                # Withholding bucket snapshots
                for bucket in withholding_buckets:
                    contrib = transfer_entries.filter(
                        withholding_category=bucket,
                        to_account=bucket.account,
                    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                    MonthEndWithholdingCategorySnapshot.objects.create(
                        month_close=month_close,
                        withholding_category=bucket,
                        monthly_target=bucket.monthly_target or Decimal('0'),
                        actual_contributed=contrib,
                    )

                # Income category snapshots
                for inc_cat in IncomeCategory.objects.filter(monthly_target__gt=0):
                    received = income_entries.filter(
                        income_category=inc_cat
                    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                    MonthEndIncomeCategorySnapshot.objects.create(
                        month_close=month_close,
                        income_category=inc_cat,
                        monthly_target=inc_cat.monthly_target,
                        actual_received=received,
                    )

            # Success message with backup details
            print(f"[MONTH-END] Saving account and net worth snapshots...")
            backup_size_mb = backup_info['zip_size'] / 1024 / 1024
            component_count = len(backup_info['components'])

            print(f"\n{'='*60}")
            print(f"[MONTH-END] ✅ Month-end close completed successfully!")
            print(f"[MONTH-END] Total time: see timestamps above")
            print(f"{'='*60}\n")

            messages.success(
                request,
                f'✅ {month_display} closed successfully! '
                f'Backup: {backup_filename} ({backup_size_mb:.1f} MB, {component_count} components)'
            )
            return redirect('dashboard')

        except Exception as e:
            messages.error(request, f'Error closing month: {str(e)}')
            return redirect('month_end_wizard')

    # Step 5 GET: Show confirmation
    elif step == '5':
        context = {
            'step': 5,
            'month': month_str,
            'month_display': month_display,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_savings': net_savings,
            'transaction_count': transaction_count,
        }
        return render(request, 'month_end_wizard.html', context)

    # Default: redirect to step 1
    return redirect('month_end_wizard')


def net_worth_tracker(request):
    """
    Comprehensive Net Worth tracking page showing:
    - Current assets and liabilities
    - Historical net worth progression (from month-end closes)
    - Breakdown by asset type (liquid, investment, property)
    - Property details with market value and mortgages
    """
    from .models import MonthEndClose, NetWorthSnapshot, RentalProperty, AccountSnapshot

    # ============================================
    # CURRENT NET WORTH (Real-time)
    # ============================================

    # Assets
    liquid_assets = Decimal('0')
    retirement_assets = Decimal('0')
    liquid_accounts = []
    retirement_accounts = []

    for account in BankAccount.objects.filter(is_active=True):
        balance = account.current_balance or Decimal('0')
        account_data = {
            'name': account.name,
            'institution': account.institution,
            'type': account.get_account_type_display(),
            'balance': balance
        }

        if account.account_type == 'RETIREMENT':
            retirement_assets += balance
            retirement_accounts.append(account_data)
        else:
            # TFSA is liquid (can withdraw anytime), along with chequing, savings
            liquid_assets += balance
            liquid_accounts.append(account_data)

    # Property Assets and Liabilities
    properties = []
    total_property_value = Decimal('0')
    total_mortgages = Decimal('0')
    total_property_equity = Decimal('0')

    for prop in RentalProperty.objects.filter(is_active=True):
        estimated_value = prop.estimated_value or Decimal('0')
        mortgage_balance = prop.total_mortgage_balance or Decimal('0')
        equity = prop.equity or Decimal('0')

        properties.append({
            'name': prop.name,
            'market_value': estimated_value,
            'mortgage_balance': mortgage_balance,
            'equity': equity,
            'notes': prop.notes or ''
        })

        total_property_value += estimated_value
        total_mortgages += mortgage_balance
        total_property_equity += equity

    # Credit Card Liabilities
    credit_card_debt = Decimal('0')
    credit_cards = []

    for account in BankAccount.objects.filter(is_active=True, account_type='CREDIT_CARD'):
        balance = account.current_balance or Decimal('0')
        if balance < 0:  # Credit cards show negative balance as debt
            debt = abs(balance)
            credit_card_debt += debt
            credit_cards.append({
                'name': account.name,
                'balance': debt
            })

    # Total Current Net Worth
    total_assets = liquid_assets + retirement_assets + total_property_value
    total_liabilities = total_mortgages + credit_card_debt
    current_net_worth = total_assets - total_liabilities

    # ============================================
    # HISTORICAL NET WORTH (From Month-End Closes)
    # ============================================

    month_closes = MonthEndClose.objects.filter(
        is_locked=True
    ).order_by('month')

    historical_data = []
    prev_net_worth = None
    for close in month_closes:
        # Get the associated net worth snapshot
        snapshot = NetWorthSnapshot.objects.filter(month_close=close).first()

        if snapshot:
            # Calculate change from previous month
            change = None
            if prev_net_worth is not None:
                change = snapshot.total_net_worth - prev_net_worth

            historical_data.append({
                'month': close.month,
                'month_display': close.month.strftime('%b %Y'),
                'liquid_assets': snapshot.liquid_assets,
                'investment_assets': snapshot.investment_assets,
                'property_value': snapshot.property_value,
                'liabilities': snapshot.liabilities,
                'total_net_worth': snapshot.total_net_worth,
                'change': change
            })

            prev_net_worth = snapshot.total_net_worth

    # Calculate month-over-month changes
    month_over_month_change = Decimal('0')
    month_over_month_change_abs = Decimal('0')
    month_over_month_percent = Decimal('0')
    latest_month_display = None
    previous_month_display = None
    latest_net_worth = None
    previous_net_worth = None
    year_over_year_change = Decimal('0')
    year_over_year_change_abs = Decimal('0')
    year_over_year_percent = Decimal('0')

    if len(historical_data) >= 2:
        latest = historical_data[-1]
        previous = historical_data[-2]
        month_over_month_change = latest['total_net_worth'] - previous['total_net_worth']
        month_over_month_change_abs = abs(month_over_month_change)
        latest_month_display = latest['month_display']
        previous_month_display = previous['month_display']
        latest_net_worth = latest['total_net_worth']
        previous_net_worth = previous['total_net_worth']
        if previous['total_net_worth'] != 0:
            month_over_month_percent = (month_over_month_change / previous['total_net_worth']) * 100

    if len(historical_data) >= 13:
        latest = historical_data[-1]
        year_ago = historical_data[-13]
        year_over_year_change = latest['total_net_worth'] - year_ago['total_net_worth']
        year_over_year_change_abs = abs(year_over_year_change)
        if year_ago['total_net_worth'] != 0:
            year_over_year_percent = (year_over_year_change / year_ago['total_net_worth']) * 100

    # Prepare chart data (JSON serializable)
    chart_labels = [item['month_display'] for item in historical_data]
    chart_net_worth = [float(item['total_net_worth']) for item in historical_data]
    chart_liquid = [float(item['liquid_assets']) for item in historical_data]
    chart_investment = [float(item['investment_assets']) for item in historical_data]
    chart_property = [float(item['property_value']) for item in historical_data]
    chart_liabilities = [float(item['liabilities']) for item in historical_data]

    # Asset Allocation (Current)
    total_assets_float = float(total_assets) if total_assets > 0 else 1
    liquid_percent = (float(liquid_assets) / total_assets_float) * 100
    retirement_percent = (float(retirement_assets) / total_assets_float) * 100
    property_percent = (float(total_property_value) / total_assets_float) * 100

    context = {
        # Current Net Worth
        'current_net_worth': current_net_worth,
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,

        # Assets Breakdown
        'liquid_assets': liquid_assets,
        'retirement_assets': retirement_assets,
        'total_property_value': total_property_value,
        'liquid_accounts': liquid_accounts,
        'retirement_accounts': retirement_accounts,

        # Liabilities Breakdown
        'total_mortgages': total_mortgages,
        'credit_card_debt': credit_card_debt,
        'properties': properties,
        'credit_cards': credit_cards,

        # Asset Allocation
        'liquid_percent': liquid_percent,
        'retirement_percent': retirement_percent,
        'property_percent': property_percent,

        # Historical Data
        'historical_data': historical_data,
        'month_over_month_change': month_over_month_change,
        'month_over_month_change_abs': month_over_month_change_abs,
        'month_over_month_percent': month_over_month_percent,
        'latest_month_display': latest_month_display,
        'previous_month_display': previous_month_display,
        'latest_net_worth': latest_net_worth,
        'previous_net_worth': previous_net_worth,
        'year_over_year_change': year_over_year_change,
        'year_over_year_change_abs': year_over_year_change_abs,
        'year_over_year_percent': year_over_year_percent,

        # Chart Data (for JavaScript)
        'chart_labels': chart_labels,
        'chart_net_worth': chart_net_worth,
        'chart_liquid': chart_liquid,
        'chart_investment': chart_investment,
        'chart_property': chart_property,
        'chart_liabilities': chart_liabilities,
    }

    return render(request, 'net_worth_tracker.html', context)


import base64
import json
import webauthn
from webauthn.helpers.structs import (
    AuthenticatorSelectionCriteria,
    UserVerificationRequirement,
    ResidentKeyRequirement,
)
from webauthn.helpers import base64url_to_bytes
from django.contrib.auth import login as auth_login
from .models import WebAuthnCredential


def _webauthn_rp_id():
    return settings.WEBAUTHN_RP_ID


def _webauthn_origin():
    return settings.WEBAUTHN_ORIGIN


@require_POST
def webauthn_register_begin(request):
    user = request.user
    options = webauthn.generate_registration_options(
        rp_id=_webauthn_rp_id(),
        rp_name=settings.WEBAUTHN_RP_NAME,
        user_id=str(user.id).encode(),
        user_name=user.username,
        user_display_name=user.get_full_name() or user.username,
        authenticator_selection=AuthenticatorSelectionCriteria(
            resident_key=ResidentKeyRequirement.REQUIRED,
            user_verification=UserVerificationRequirement.REQUIRED,
        ),
    )
    request.session["webauthn_register_challenge"] = base64.b64encode(options.challenge).decode()
    return JsonResponse(json.loads(webauthn.options_to_json(options)))


@require_POST
def webauthn_register_complete(request):
    try:
        data = json.loads(request.body)
        challenge = base64.b64decode(request.session.pop("webauthn_register_challenge", ""))
        verification = webauthn.verify_registration_response(
            credential=data,
            expected_challenge=challenge,
            expected_rp_id=_webauthn_rp_id(),
            expected_origin=_webauthn_origin(),
        )
        WebAuthnCredential.objects.create(
            user=request.user,
            credential_id=bytes(verification.credential_id),
            public_key=bytes(verification.credential_public_key),
            sign_count=verification.sign_count,
            device_name=data.get("deviceName", "Passkey"),
        )
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


@require_POST
def webauthn_auth_begin(request):
    options = webauthn.generate_authentication_options(
        rp_id=_webauthn_rp_id(),
        user_verification=UserVerificationRequirement.REQUIRED,
    )
    request.session["webauthn_auth_challenge"] = base64.b64encode(options.challenge).decode()
    return JsonResponse(json.loads(webauthn.options_to_json(options)))


@require_POST
def webauthn_auth_complete(request):
    try:
        data = json.loads(request.body)
        challenge = base64.b64decode(request.session.pop("webauthn_auth_challenge", ""))
        raw_id = base64url_to_bytes(data["rawId"])
        cred = WebAuthnCredential.objects.get(credential_id=raw_id)
        verification = webauthn.verify_authentication_response(
            credential=data,
            expected_challenge=challenge,
            expected_rp_id=_webauthn_rp_id(),
            expected_origin=_webauthn_origin(),
            credential_public_key=bytes(cred.public_key),
            credential_current_sign_count=cred.sign_count,
        )
        cred.sign_count = verification.new_sign_count
        cred.save(update_fields=["sign_count"])
        auth_login(request, cred.user, backend="django.contrib.auth.backends.ModelBackend")
        return JsonResponse({"ok": True})
    except WebAuthnCredential.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Passkey not recognised."}, status=400)
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=400)


@require_POST
def webauthn_delete(request):
    cred_id = request.POST.get("credential_id")
    WebAuthnCredential.objects.filter(user=request.user, pk=cred_id).delete()
    return redirect("profile")


def profile(request):
    user = request.user
    if request.method == "POST":
        user.first_name = request.POST.get("first_name", "").strip()
        user.last_name = request.POST.get("last_name", "").strip()
        user.email = request.POST.get("email", "").strip()
        user.save()
        messages.success(request, "Profile updated.")
        return redirect("profile")
    credentials = WebAuthnCredential.objects.filter(user=user).order_by("-created_at")
    return render(request, "profile.html", {"user": user, "credentials": credentials})

